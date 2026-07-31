# -*- coding: utf-8 -*-
"""
ZETA ENTERPRISE PLATFORM — sync_users.py
=====================================================================
Automated, idempotent synchronization utility.
Reads active Line Managers from the Zeta Commercial Database, maps them to 
canonical business units and lines, updates the user access configuration 
spreadsheet with secure credentials, generates the client-side permission 
and authentication caches, runs post-build validation, and executes 
JSDOM functional tests. Includes automatic transactional rollback.
=====================================================================
"""
import os
import sys
import shutil
import hashlib
import json
import time
import subprocess
from datetime import datetime
import openpyxl

# Force stdout to UTF-8 to prevent console encoding crashes on Windows
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# Constants
SALT = "ZETA2026INTEL"
COMM_DB = "COMMERCCIAL DATA BASE.xlsx"
USER_CONFIG = "Zeta_Dashboard_User_Config.xlsx"
BACKUP_DIR = "backups"

SECURITY_CACHE = "cache/security-cache.json"
PERMISSIONS_CACHE = "cache/permissions-cache.json"
USER_AUTH_JS = "cache/userAuth.js"

IQVIA_DATA_JS = "cache/iqvia.data.js"
IQVIA_JSON = "cache/iqvia.json"

TEST_CREDENTIALS = "test_credentials.json"
REPORT_FILE = "Sync_Report.txt"

# Canonical Mappings
BU_MAP = {
    'DIABETES': 'DIAB',
    'DIAB': 'DIAB',
    'GIT': 'GIT',
    'CLUSTER': 'Cluster',
    'CHC': 'CHC'
}

LINE_MAP = {
    'GIT I': 'GIT-I',
    'GIT II': 'GIT-II',
    'GIT III': 'GIT-III',
    'CNS': 'CNS',
    'DERMA': 'Derma',
    'CVM I': 'CVM-I',
    'CVM-I': 'CVM-I',
    'CVM II': 'CVM-II',
    'CVM-II': 'CVM-II',
    'ORTHO I': 'ORTHO-I',
    'ORTHO-I': 'ORTHO-I',
    'ORTHO II': 'ORTHO-II',
    'ORTHO-II': 'ORTHO-II',
    'PEDIA/GYN': 'PEDIA',
    'PEDIA': 'PEDIA',
    'DIABETES I': 'DIAB-I',
    'DIAB I': 'DIAB-I',
    'DIAB-I': 'DIAB-I',
    'DIABETES II': 'DIAB-II',
    'DIAB II': 'DIAB-II',
    'DIAB-II': 'DIAB-II',
    'DIABETES III': 'DIAB-III',
    'DIAB III': 'DIAB-III',
    'DIAB-III': 'DIAB-III',
    'DIABETES IV': 'DIAB-IV',
    'DIAB IV': 'DIAB-IV',
    'DIAB-IV': 'DIAB-IV',
    'CHC': 'CHC'
}

def clean_str(v):
    if v is None: return ""
    return str(v).strip()

def normalize_bu(bu):
    clean_bu = clean_str(bu).upper()
    for k, v in BU_MAP.items():
        if k in clean_bu:
            return v
    return clean_str(bu)

def normalize_line(line):
    clean_line = clean_str(line).upper().replace('_', '-').replace('/', '-').strip()
    # Resolve aliases (sort descending by key length to avoid substring collision)
    for k, v in sorted(LINE_MAP.items(), key=lambda x: -len(x[0])):
        if k in clean_line:
            return v
    # Default fallback title casing
    return clean_str(line).title()

def sha256_hex(s):
    return hashlib.sha256(s.encode('utf-8')).hexdigest()

def make_backup():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    folder = os.path.join(ROOT_DIR, BACKUP_DIR, f"backup_{timestamp}")
    os.makedirs(folder, exist_ok=True)
    
    files_to_backup = [
        (USER_CONFIG, "root_user_config.xlsx"),
        (os.path.join("iqvia_source", "config", "Zeta_Dashboard_User_Config.xlsx"), "iqvia_user_config.xlsx"),
        (SECURITY_CACHE, "security_cache.json"),
        (PERMISSIONS_CACHE, "permissions_cache.json"),
        (USER_AUTH_JS, "userAuth.js"),
        (IQVIA_DATA_JS, "iqvia_data.js"),
        (IQVIA_JSON, "iqvia.json")
    ]
    backed_up = []
    for rel_path, bkp_name in files_to_backup:
        src = os.path.join(ROOT_DIR, rel_path)
        if os.path.exists(src):
            shutil.copy2(src, os.path.join(folder, bkp_name))
            backed_up.append(rel_path)
    print(f"[Backup] Saved backup of {len(backed_up)} files to: {folder}")
    return folder

def restore_backup(folder):
    print(f"[Rollback] Restoring original configuration from: {folder} ...")
    restore_map = [
        ("root_user_config.xlsx", USER_CONFIG),
        ("iqvia_user_config.xlsx", os.path.join("iqvia_source", "config", "Zeta_Dashboard_User_Config.xlsx")),
        ("security_cache.json", SECURITY_CACHE),
        ("permissions_cache.json", PERMISSIONS_CACHE),
        ("userAuth.js", USER_AUTH_JS),
        ("iqvia_data.js", IQVIA_DATA_JS),
        ("iqvia.json", IQVIA_JSON)
    ]
    for bkp_name, rel_path in restore_map:
        src = os.path.join(folder, bkp_name)
        if os.path.exists(src):
            dest = os.path.join(ROOT_DIR, rel_path)
            os.makedirs(os.path.dirname(dest), exist_ok=True)
            shutil.copy2(src, dest)
    print("[Rollback] Restore complete.")

def main():
    t_start = time.time()
    print("=== STARTING USER SYNCHRONIZATION PROCESS ===")
    
    # 1. Backups
    backup_folder = make_backup()
    
    users_created = 0
    users_updated = 0
    users_skipped = 0
    duplicates_ignored = 0
    inactive_ignored = 0
    
    try:
        # 2. Parse Zeta Commercial Database
        print(f"[Database] Loading {COMM_DB} ...")
        db_wb = openpyxl.load_workbook(os.path.join(ROOT_DIR, COMM_DB), data_only=True)
        db_sheet = db_wb.active
        db_rows = list(db_sheet.iter_rows(values_only=True))
        db_headers = [clean_str(h) for h in db_rows[0]]
        db_ci = {h: i for i, h in enumerate(db_headers)}
        
        # Managers to extract
        scanned_managers = {}
        for row in db_rows[1:]:
            if not row or not any(row): continue
            
            # Extract core fields
            email = clean_str(row[db_ci.get('Business Email')]).strip().lower()
            name = clean_str(row[db_ci.get('Employee Name (English)')]).strip()
            code = clean_str(row[db_ci.get('Code')]).strip()
            pos = clean_str(row[db_ci.get('Position (English)')]).strip()
            status = clean_str(row[db_ci.get('Status')]).strip()
            bu_raw = clean_str(row[db_ci.get('Business Unit')]).strip()
            line_raw = clean_str(row[db_ci.get('Line')]).strip()
            
            if not email or not name:
                continue
                
            # Filter inactive or vacant
            if status.lower() != 'active':
                inactive_ignored += 1
                continue
            if 'vacant' in name.lower() or 'vacant' in email:
                continue
                
            # Filter positions matching "Line Manager"
            pos_lower = pos.lower()
            is_line_mgr = ('national sales manager' in pos_lower or 
                           'line sales manager' in pos_lower)
                           
            if not is_line_mgr:
                continue
                
            # Deduplicate by email and code
            unique_key = email
            if unique_key in scanned_managers:
                duplicates_ignored += 1
                continue
                
            # Normalize BU and Line
            norm_bu = normalize_bu(bu_raw)
            norm_line = normalize_line(line_raw)
            
            scanned_managers[unique_key] = {
                'name': name,
                'email': email,
                'code': code,
                'position': pos,
                'bu': norm_bu,
                'line': norm_line
            }
        
        db_wb.close()
        print(f"[Database] Found {len(scanned_managers)} active Line Managers in source sheet.")

        # 3. Read Destination Workbook Config
        print(f"[Config] Loading {USER_CONFIG} ...")
        cfg_wb = openpyxl.load_workbook(os.path.join(ROOT_DIR, USER_CONFIG))
        cfg_sheet = cfg_wb['Users']
        cfg_rows = list(cfg_sheet.iter_rows(values_only=True))
        
        # Locate header row
        header_idx = None
        for i, row in enumerate(cfg_rows):
            if row and 'Email' in [clean_str(c) for c in row if c]:
                header_idx = i
                break
                
        if header_idx is None:
            raise ValueError("Could not locate headers in Users sheet of configuration file.")
            
        headers = [clean_str(c) for c in cfg_rows[header_idx]]
        ci = {h: i for i, h in enumerate(headers)}
        
        # Parse existing users
        existing_users = []
        for i, row in enumerate(cfg_rows[header_idx + 1:], start=header_idx + 1):
            if not row or not row[ci.get('Email')]: continue
            u_email = clean_str(row[ci['Email']]).strip().lower()
            existing_users.append({
                'row_num': i + 1,
                'email': u_email,
                'name': clean_str(row[ci['Full Name']]),
                'pwd': clean_str(row[ci['Password']]),
                'role': clean_str(row[ci['Role']]),
                'bu': clean_str(row[ci['Allowed BU\n(comma-separated or ALL)']]),
                'lines': clean_str(row[ci['Allowed Lines\n(comma-separated or ALL)']]),
                'dm1': clean_str(row[ci['Allowed Markets DM1\n(comma-separated or ALL)']]),
                'prods': clean_str(row[ci['Allowed Products\n(comma-separated or ALL)']]),
                'export': clean_str(row[ci['Can Export\n(Yes/No)']]),
                'active': clean_str(row[ci['Active\n(Yes/No)']]),
                'notes': clean_str(row[ci['Notes']])
            })
            
        print(f"[Config] Found {len(existing_users)} existing user records.")
        
        # Identify first name frequencies to check for duplicates
        first_names = []
        for u in existing_users:
            if u['name']:
                first_names.append(u['name'].strip().split()[0].capitalize())
        for u_email, m in scanned_managers.items():
            if m['name']:
                first_names.append(m['name'].strip().split()[0].capitalize())
                
        # Merge & Update logic
        final_users = []
        existing_emails = {u['email']: u for u in existing_users}
        
        # Process existing accounts (preserve admin / BU Managers / custom roles)
        for u in existing_users:
            email = u['email']
            # If they are a Line Manager in the database, we update their details
            if email in scanned_managers:
                m = scanned_managers[email]
                # Update properties if and only if role is Line Manager
                if u['role'] == 'Line Manager' or u['role'] == '':
                    u['name'] = m['name']
                    u['role'] = 'Line Manager'
                    u['bu'] = m['bu']
                    u['lines'] = m['line']
                    u['active'] = 'Yes'
                    u['notes'] = m['position']
                    users_updated += 1
                final_users.append(u)
            else:
                # Keep unchanged (SFE Manager, BEX, CEO, VP, Admin, BU Managers etc)
                final_users.append(u)
                
        # Add brand new Line Managers
        for email, m in scanned_managers.items():
            if email not in existing_emails:
                first_name = m['name'].strip().split()[0].capitalize()
                # Check for first name duplicate
                if first_names.count(first_name) > 1:
                    pwd = f"{first_name}{m['code']}@2026"
                else:
                    pwd = f"{first_name}@2026"
                    
                new_u = {
                    'email': email,
                    'name': m['name'],
                    'pwd': pwd,
                    'role': 'Line Manager',
                    'bu': m['bu'],
                    'lines': m['line'],
                    'dm1': 'ALL',
                    'prods': 'ALL',
                    'export': 'Yes',
                    'active': 'Yes',
                    'notes': m['position']
                }
                final_users.append(new_u)
                users_created += 1

        # Clear and rewrite destination sheet
        # Preserve formatting of the title rows (first 3 rows) by only deleting from header_idx + 1 onward
        max_row = cfg_sheet.max_row
        if max_row > header_idx + 1:
            cfg_sheet.delete_rows(header_idx + 2, max_row - header_idx)
            
        # Write users
        for idx, u in enumerate(final_users, start=1):
            row_vals = [None] * len(headers)
            row_vals[ci['#']] = idx
            row_vals[ci['Full Name']] = u['name']
            row_vals[ci['Email']] = u['email']
            row_vals[ci['Password']] = u['pwd']
            row_vals[ci['Role']] = u['role']
            row_vals[ci['Allowed BU\n(comma-separated or ALL)']] = u['bu']
            row_vals[ci['Allowed Lines\n(comma-separated or ALL)']] = u['lines']
            row_vals[ci['Allowed Markets DM1\n(comma-separated or ALL)']] = u['dm1']
            row_vals[ci['Allowed Products\n(comma-separated or ALL)']] = u['prods']
            row_vals[ci['Can Export\n(Yes/No)']] = u['export']
            row_vals[ci['Active\n(Yes/No)']] = u['active']
            row_vals[ci['Notes']] = u['notes']
            
            cfg_sheet.append(row_vals)
            
        cfg_wb.save(os.path.join(ROOT_DIR, USER_CONFIG))
        cfg_wb.close()
        print(f"[Config] Saved updated user configuration sheet successfully. Created: {users_created}, Updated: {users_updated}")

        # Copy to iqvia_source/config/ to ensure refresh_iqvia.py reads it
        dest_config = os.path.join(ROOT_DIR, 'iqvia_source', 'config', 'Zeta_Dashboard_User_Config.xlsx')
        os.makedirs(os.path.dirname(dest_config), exist_ok=True)
        shutil.copy2(os.path.join(ROOT_DIR, USER_CONFIG), dest_config)
        print(f"[Config] Copied updated config to: {dest_config}")

        # 4. Generate Caches (security-cache, permissions-cache, userAuth.js)
        print("[Cache] Rebuilding security and permissions caches...")
        
        security_users = {}
        permissions = {}
        user_auth_dict = {}
        
        for u in final_users:
            if u['active'].upper() != 'YES':
                continue
            email = u['email']
            pwd = u['pwd']
            
            # Hashing
            pwd_hash = sha256_hex(email + ":" + pwd + ":" + SALT)
            
            # Format BU/Lines as arrays or null if ALL
            def parse_perms(val):
                if not val or val.strip().upper() == 'ALL':
                    return None
                return [x.strip() for x in val.split(',') if x.strip()]
                
            bu_arr = parse_perms(u['bu'])
            lines_arr = parse_perms(u['lines'])
            
            security_users[email] = {
                'name': u['name'],
                'hash': pwd_hash,
                'role': u['role'],
                'bu': bu_arr,
                'lines': lines_arr
            }
            
            permissions[email] = {
                'bu': bu_arr,
                'lines': lines_arr,
                'role': u['role'],
                'canExport': u['export'].upper() == 'YES'
            }
            
            user_auth_dict[email] = {
                'name': u['name'],
                'role': u['role'],
                'bu': bu_arr,
                'lines': lines_arr,
                'canExport': u['export'].upper() == 'YES'
            }
            
        # Write JSON files (deterministic ordering)
        os.makedirs(os.path.join(ROOT_DIR, "cache"), exist_ok=True)
        
        # Sort users by email key
        sorted_sec_users = {k: security_users[k] for k in sorted(security_users.keys())}
        sorted_perms = {k: permissions[k] for k in sorted(permissions.keys())}
        sorted_user_auth = {k: user_auth_dict[k] for k in sorted(user_auth_dict.keys())}
        
        with open(os.path.join(ROOT_DIR, SECURITY_CACHE), 'w', encoding='utf-8') as f:
            json.dump({'users': sorted_sec_users}, f, indent=2, sort_keys=True)
            
        with open(os.path.join(ROOT_DIR, PERMISSIONS_CACHE), 'w', encoding='utf-8') as f:
            json.dump({'permissions': sorted_perms}, f, indent=2, sort_keys=True)
            
        with open(os.path.join(ROOT_DIR, USER_AUTH_JS), 'w', encoding='utf-8') as f:
            f.write(f"window.USER_AUTH = {json.dumps(sorted_user_auth, indent=2, sort_keys=True)};\n")
            
        print("[Cache] Generated security-cache.json, permissions-cache.json, and userAuth.js successfully.")

        # 5. Run IQVIA Security Cache Refresh
        print("[ETL] Rebuilding main dashboard user database by running refresh_iqvia.py ...")
        ref_res = subprocess.run([sys.executable, "refresh_iqvia.py"], cwd=ROOT_DIR, capture_output=True, text=True)
        if ref_res.returncode != 0:
            raise RuntimeError(f"refresh_iqvia.py execution failed: {ref_res.stderr}")
        print("[ETL] refresh_iqvia.py executed successfully.")

        # 6. Write Test Credentials for the test runner
        # Find representative CHC, GIT, Cluster users for testing
        test_creds = {
            'admin': {'email': 'ahmed.abdullah@zeta-pharma.com', 'password': 'Zeta@2026'},
            'chc': None,
            'git': None,
            'cluster': None
        }
        
        for u in final_users:
            email = u['email']
            role = u['role']
            bu = u['bu']
            pwd = u['pwd']
            if role == 'Line Manager':
                if bu == 'CHC' and not test_creds['chc']:
                    test_creds['chc'] = {'email': email, 'password': pwd}
                if bu == 'GIT' and not test_creds['git']:
                    test_creds['git'] = {'email': email, 'password': pwd}
                if bu == 'Cluster' and not test_creds['cluster']:
                    test_creds['cluster'] = {'email': email, 'password': pwd}
                    
        # Fallbacks if some BU wasn't synced
        if not test_creds['chc']: test_creds['chc'] = {'email': 'amr.khalifa@zeta-pharma.com', 'password': 'Amr@2026'}
        if not test_creds['git']: test_creds['git'] = {'email': 'nader.khaled@zeta-pharma.com', 'password': 'Nader473@2026'}
        if not test_creds['cluster']: test_creds['cluster'] = {'email': 'mohamed.elkerdawy@zeta-pharma.com', 'password': 'Mohamed86@2026'}
        
        with open(os.path.join(ROOT_DIR, TEST_CREDENTIALS), 'w', encoding='utf-8') as f:
            json.dump(test_creds, f, indent=2)
        print("[Test] Wrote test_credentials.json successfully.")

        # 7. Run Automatic Validation
        print("[Validation] Running sanity checks...")
        
        # Verify active line managers exist in security-cache
        for email, m in scanned_managers.items():
            if email not in security_users:
                raise ValueError(f"Validation Error: Synced manager {email} is missing from active security cache.")
                
        # Verify admin accounts preserved
        for admin_email in ['ahmed.abdullah@zeta-pharma.com', 'ahmed.hamid@zeta-pharma.com', 'mohammed.bakr@zeta-pharma.com']:
            if admin_email not in security_users:
                raise ValueError(f"Validation Error: Admin account {admin_email} was removed or disabled.")
            if security_users[admin_email]['bu'] is not None:
                raise ValueError(f"Validation Error: Admin account {admin_email} permissions were restricted.")

        # Verify role, BU and line structure
        for email, u in security_users.items():
            if not u['role']:
                raise ValueError(f"Validation Error: User {email} has no assigned role.")
            # Line Managers must have a single BU and Line (or array of lines)
            if u['role'] == 'Line Manager':
                if not u['bu'] or len(u['bu']) != 1:
                    raise ValueError(f"Validation Error: Line Manager {email} has invalid BU assignment: {u['bu']}")
                if not u['lines'] or len(u['lines']) == 0:
                    raise ValueError(f"Validation Error: Line Manager {email} has no line assignment.")
                    
        # Verify file existence
        for path_to_check in [SECURITY_CACHE, PERMISSIONS_CACHE, USER_AUTH_JS, IQVIA_DATA_JS]:
            if not os.path.exists(os.path.join(ROOT_DIR, path_to_check)):
                raise ValueError(f"Validation Error: Cache file {path_to_check} is missing.")
                
        print("[Validation] Core sanity checks passed.")

        # 8. Run Functional Test Suite
        print("[Test] Launching automated functional tests inside JSDOM...")
        test_res = subprocess.run(["node", "test_security_auth.js"], cwd=ROOT_DIR, capture_output=True, text=True)
        print(test_res.stdout)
        
        if test_res.returncode != 0:
            print(test_res.stderr, file=sys.stderr)
            raise RuntimeError("Functional security test suite failed.")
            
        print("[Test] Automated functional tests completed successfully.")

        # 9. Sync Report Generation
        duration = time.time() - t_start
        report_text = f"""ZETA USER SYNCHRONIZATION REPORT
=========================================
Timestamp: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Duration: {duration:.2f} seconds
Result: SUCCESS

Users Created: {users_created}
Users Updated: {users_updated}
Users Skipped (Inactive): {users_skipped}
Duplicates Ignored: {duplicates_ignored}
Inactive Ignored in Source: {inactive_ignored}

VALIDATION RESULTS:
-----------------------------
- Active managers synced: YES
- No duplicate accounts: YES
- Admin accounts preserved: YES
- Deterministic cache ordering: YES
- JSDOM functional tests: PASSED

SYSTEM STATUS:
-----------------------------
- Zeta_Dashboard_User_Config.xlsx updated.
- security-cache.json generated.
- permissions-cache.json generated.
- userAuth.js generated.
- iqvia.data.js regenerated.
"""
        with open(os.path.join(ROOT_DIR, REPORT_FILE), 'w', encoding='utf-8') as f:
            f.write(report_text)
            
        print("=== SYNCHRONIZATION COMPLETED SUCCESSFULLY ===")
        
    except Exception as err:
        print(f"\n[CRITICAL ERROR] Synchronization Failed: {err}", file=sys.stderr)
        # Roll back changes
        restore_backup(backup_folder)
        
        # Write error report
        duration = time.time() - t_start
        error_report = f"""ZETA USER SYNCHRONIZATION REPORT
=========================================
Timestamp: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
Duration: {duration:.2f} seconds
Result: FAILED

Error Details: {err}

ROLLBACK PERFORMED: YES (All files restored to original values)
"""
        with open(os.path.join(ROOT_DIR, REPORT_FILE), 'w', encoding='utf-8') as f:
            f.write(error_report)
            
        sys.exit(1)

if __name__ == '__main__':
    # Set root directory relative to this script
    ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    main()
