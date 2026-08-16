"""
etl/build_sprint_cache.py
==========================
Builds cache/sprint.data.js -- the Zeta Sprint 2026 monthly competition
standings (Medical Rep, CHC Sales Rep, DM/DSM, ASM/NSM tiers).

Mirrors the other etl/build_*_cache.py scripts: reads source workbooks +
the existing Coverage/Sales caches, computes a self-contained JSON payload,
gzip+base64 encodes it, and writes it as `window.SPRINT_CACHE = {b64Data:"..."}`
so js/sprint.js can load it exactly like every other page's cache.

METHODOLOGY (confirmed with Ahmed, 2026-08-15):
  - Ranking is PER MONTH, not a rolling "as of today" snapshot. This build
    produces the June 2026 ranking; re-run monthly as new periods close.
  - Probation: hire day 1-15 -> reference = 1st of same month; day 16-31
    -> reference = 1st of next month; probation passes 3 months later.
    A rep is eligible for month M's ranking only if their probation-passed
    date is on/before the 1st of M. Verified against the Coverage cache's
    own period-level "Non-Probation" flag: zero disagreements (837
    employees x 5 periods).
  - Active/not-resigned: sourced from Database Shortcut.xlsx (the HR
    database) by employee Code -- never by name. Uses Last Day of Work as
    the ground truth for period membership (Status alone lags real
    resignations by weeks in ~36 cases company-wide); Status is only a
    fallback when no Last Day is on file.
  - Curve sheet: "medical rep points scheme" / "SALES REP" in
    scaling scores.xlsx -- confirmed final (only sheet whose point caps
    sum to the deck's stated 100).
  - CHC/Sales Rep Coverage: sheet's Coverage curve tops out at 10 pts;
    scaled x4 to hit the deck's stated 40pt weight -- confirmed approach
    pending workbook-author verification.
  - DM/DSM: Team Avg component (70 of 100), computed as the mean June
    totalPts of eligible reps directly under each DM/DSM, using the SAME
    hierarchy fields already in the Coverage cache.
  - ASM/NSM: Team Avg component (80 of 100) -- per Ahmed 2026-08-15
    ("ASM and NSM average points for their DSM"), computed as the mean
    totalPts of the DM/DSMs reporting up to that ASM/NSM (NOT individual
    reps directly), matching the real org chain Rep -> DM/DSM -> ASM/NSM.
    Each DM/DSM's ASM/NSM is derived by majority vote over their own
    reps' direct ASM/NSM field in the Coverage cache.
  - Field Working Days / DV Coverage / Calls-per-DV are NOT computable
    from any existing source -- Ahmed provides a sheet (see
    TEMPLATE_SHEETS). Those KPI slots are emitted as null
    (pendingDataFeed=true), never silently scored as zero.
  - Brand Manager: NOT built yet -- Ahmed will provide data later.
  - WINNER FLOOR, extended to DM/DSM/ASM/NSM 2026-08-16 (see js/sprint.js
    module doc for the full rule and UI): every manager tier now carries
    teamSalesVal/teamSalesTgt/teamSalesAchPct, computed as SUM(each
    eligible rep's raw sales value)/SUM(their raw sales target) across
    the WHOLE reporting subtree beneath that manager, divided exactly
    ONCE at output -- never an average of already-divided percentages at
    any tier. DM/DSM sums its own reps directly; ASM/NSM sums its DM/DSMs'
    already-team-summed val/tgt (propagated via the ASM/NSM pool build),
    so the ratio always traces back to raw rep-level sales no matter how
    many hierarchy levels up. js/sprint.js gates the 🏆 WINNER/🥈
    RUNNER-UP and (DM/DSM only) 🎖 BU Leader badges plus the cash payout
    on teamSalesAchPct >= 70%, cascading to the next eligible manager in
    rank order exactly like the rep-level floor -- never re-sorting the
    honest ranking itself. If nobody in a scope (Line for reps, BU for
    DM/DSM, company-wide for ASM/NSM) clears 70%, that scope has no
    winner and no payout that month.
"""
import re
import os
import base64
import gzip
import json
import time
import datetime
from collections import defaultdict, Counter

import openpyxl

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CACHE_DIR = os.path.join(ROOT_DIR, 'cache')
DB_PATH = os.path.join(ROOT_DIR, 'Database Shortcut.xlsx')
SCALING_PATH = os.path.join(ROOT_DIR, 'zeta sprint', 'scaling scores.xlsx')
TEMPLATE_PATH = os.path.join(ROOT_DIR, 'zeta sprint', 'Sprint_Missing_KPI_Template.xlsx')
OUT_JS = os.path.join(CACHE_DIR, 'sprint.data.js')
OUT_JSON = os.path.join(CACHE_DIR, 'sprint.json')
HISTORY_DIR = os.path.join(CACHE_DIR, 'sprint_history')
HISTORY_INDEX_JS = os.path.join(HISTORY_DIR, 'index.js')

SCHEMA_VERSION = 9

# ---------------------------------------------------------------------
# KPI slots not computable from any existing cache -- Ahmed fills these
# into TEMPLATE_PATH (auto-generated the first time this script runs
# with no existing template; never overwritten after that, so his fills
# survive every future re-run).
#
# As of 2026-08-15: each column asks for the REAL ACHIEVEMENT metric
# (a raw fraction, e.g. 0.85 for 85%; Region Coverage asks for a raw
# count 1-5), NOT pre-scaled points -- per Ahmed's instruction "excel
# template according to real achievement in kpi not points". This is
# possible because the actual scoring curves for these KPIs were found
# in scaling scores.xlsx's 'dsm points scheme', 'ASM&NSM SCHEME', and
# 'Brand Managers' sheets (previously overlooked -- only the two
# Medical Rep / CHC Sales Rep sheets had been read). build_sprint_cache.py
# now converts Ahmed's raw achievement number into points itself via
# those curves (see Step 3/6/7), the same way it already does for
# Sales Achievement / Right Frequency / Coverage.
#
# DM/DSM confirmed by Ahmed 2026-08-15 ("dm dsm is right for these
# parameters as in scaling and in ppt") back to its original 3-KPI
# structure: Team Avg (70) + Field Working Days (10) + DV Coverage (10)
# + Calls per DV (10). ASM/NSM stay at Team Avg (80) + Field Days (20)
# -- that was never in question, the ASM&NSM SCHEME sheet confirms it.
# ---------------------------------------------------------------------
TEMPLATE_SHEETS = {
    'DM_DSM': [
        ('fieldDays', 'Field Working Days -- Actual % Achieved (e.g. 0.85 for 85%)', 10),
        ('dvCoverage', 'DV Coverage -- Actual % Achieved (e.g. 0.90 for 90%)', 10),
        ('callsPerDv', 'Calls per DV -- Actual % of Target (e.g. 0.95 for 95%)', 10),
    ],
    'ASM': [
        ('fieldDays', 'Field Working Days -- Actual % Achieved (e.g. 0.85 for 85%)', 20),
    ],
    'NSM': [
        ('fieldDays', 'Field Working Days -- Actual % Achieved (e.g. 0.85 for 85%)', 20),
    ],
    # National Sales is no longer in this sheet -- it's auto-calculated
    # from the Sales cache per the Brand Manager's assigned Line (see
    # Step 7), per Ahmed 2026-08-15 ("calculate bm achievement according
    # to brands they are responsible for").
    'Brand_Manager': [
        ('regionCount', 'Regions Covered -- Actual Count (1-5)', 20),
        ('tacticalPlan', 'Tactical Plan Execution -- Actual % Achieved (e.g. 0.90 for 90%)', 30),
    ],
}

# The month being ranked. Bump these three lines each time this script is
# re-run for a newly-closed period -- everything else derives from them.
EVAL_PERIOD_NAME = 'June'        # must match a label in dims['periods']
EVAL_MONTH_STR = '2026-06'       # must match a label in sales lookups['months']
EVAL_PERIOD_START = datetime.date(2026, 6, 1)
EVAL_PERIOD_END = datetime.date(2026, 6, 30)


def log(msg):
    print(f'[{time.strftime("%H:%M:%S")}] {msg}')


def load_b64(fname):
    raw = open(os.path.join(CACHE_DIR, fname), encoding='utf-8').read()
    m = re.search(r'b64Data\s*:\s*"([^"]+)"', raw)
    return json.loads(gzip.decompress(base64.b64decode(m.group(1))))


def as_date(v):
    if isinstance(v, datetime.datetime):
        return v.date()
    if isinstance(v, datetime.date) and not isinstance(v, datetime.time):
        return v
    return None


def norm_name(s):
    return str(s).upper().replace(chr(160), ' ').strip()


def probation_passed_date(hire_date):
    if hire_date is None:
        return None
    if hire_date.day <= 15:
        ref_y, ref_m = hire_date.year, hire_date.month
    else:
        ref_y, ref_m = hire_date.year, hire_date.month + 1
        if ref_m == 13:
            ref_y += 1
            ref_m = 1
    py, pm = ref_y, ref_m + 3
    while pm > 12:
        pm -= 12
        py += 1
    return datetime.date(py, pm, 1)


CANONICAL_LINE_TO_BU = {
    "CHC": "CHC", "CHC_SALES": "CHC",
    "PEDIA": "Cluster", "ORTHO-I": "Cluster", "ORTHO-II": "Cluster", "CVM-I": "Cluster", "CVM-II": "Cluster",
    "DIAB-I": "DIAB", "DIAB-II": "DIAB", "DIAB-III": "DIAB", "DIAB-IV": "DIAB",
    "Derma": "GIT", "CNS": "GIT", "GIT-I": "GIT", "GIT-II": "GIT", "GIT-III": "GIT",
}
LINE_SYNONYMS = {
    "NEUROSCIENCE": "CNS", "DERMA": "Derma", "CHC_SALES": "CHC_SALES",
    "GIT I": "GIT-I", "GIT II": "GIT-II", "GIT III": "GIT-III",
    "ORTHO I": "ORTHO-I", "ORTHO II": "ORTHO-II",
    "CVM I": "CVM-I", "CVM II": "CVM-II",
}


def normalize_line(raw):
    if raw is None:
        return None
    s = str(raw).strip()
    su = s.upper()
    if su in LINE_SYNONYMS:
        return LINE_SYNONYMS[su]
    for c in CANONICAL_LINE_TO_BU:
        if c.upper() == su:
            return c
    return s


def line_to_bu(raw):
    return CANONICAL_LINE_TO_BU.get(normalize_line(raw))


def extract_curve(ws, col_pct, col_pts, start_row=2):
    pts = []
    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, min_col=col_pct, max_col=col_pts):
        pct, val = row[0].value, row[-1].value
        if pct is None or val is None:
            continue
        pts.append((float(pct), float(val)))
    pts.sort()
    return pts


def interp(curve, x):
    if not curve:
        return None
    if x <= curve[0][0]:
        return 0.0
    if x >= curve[-1][0]:
        return curve[-1][1]
    for i in range(1, len(curve)):
        x0, y0 = curve[i - 1]
        x1, y1 = curve[i]
        if x0 <= x <= x1:
            if x1 == x0:
                return y0
            frac = (x - x0) / (x1 - x0)
            return y0 + frac * (y1 - y0)
    return curve[-1][1]


def main():
    t0 = time.time()

    if not os.path.exists(DB_PATH):
        print(f'ERROR: source not found: {DB_PATH}')
        return
    if not os.path.exists(SCALING_PATH):
        print(f'ERROR: source not found: {SCALING_PATH}')
        return

    # -----------------------------------------------------------------
    # 0. Database Shortcut.xlsx -- authoritative hire/status/resignation
    #    source, keyed by Code. Also build a Name->Code lookup for the
    #    hierarchy roles (DM/ASM/NSM), which have no Code in the Coverage
    #    cache's own dimension lists.
    # -----------------------------------------------------------------
    log('reading Database Shortcut.xlsx ...')
    wb_db = openpyxl.load_workbook(DB_PATH, data_only=True, read_only=True)
    ws_db = wb_db['Sheet1']
    code_to_hire, code_to_status, code_to_lastday, code_to_resignnotif = {}, {}, {}, {}
    code_to_position = {}
    name_to_code = {}
    # Business Email (col 24) -- lets the frontend match a logged-in
    # dashboard user (AUTH session email, from Zeta_Dashboard_User_Config.xlsx)
    # back to their own Sprint record by code, to power a personal "My
    # Performance" card -- per Ahmed 2026-08-15 ("make better page besde
    # zeta sprint"). Confirmed exact-match against the Users sheet for
    # every Sprint-participating manager checked (e.g. code 183 == NSM
    # winner Mahmoud Mokhtar's login email).
    code_to_email = {}
    # Direct/2nd/3rd Manager columns + this file's own Business Unit column
    # (cols 9, 12-17) -- covers EVERY employee code, including ASM/NSM,
    # unlike the sprint cache's teamMembers-based rollups below (those only
    # reach DM/DSM-and-below, hence ASM/NSM never had a Direct Manager
    # before) -- per Ahmed 2026-08-15 ("asm and nsm name of manager and bu
    # not present"). Used in step 7b to attach directManager/directManagerBu
    # to every scored person.
    code_to_orgchart = {}
    bm_roster = []  # (code, name, position, line) -- Brand Manager family, roster sourced by
                     # Position text since Brand Managers own products, not a rep hierarchy.
                     # `line` (Database Shortcut col 10) is the brand/line they're responsible
                     # for -- used to auto-calculate their National Sales Achievement from the
                     # Sales cache, per Ahmed 2026-08-15 ("calculate bm achievement according
                     # to brands they are responsible for").
    for r in ws_db.iter_rows(min_row=2, values_only=True):
        code, name, hire, position, status, resign_notif, last_day = r[0], r[1], r[2], r[3], r[18], r[19], r[20]
        line = r[10]
        if code is None:
            continue
        key = str(int(code)) if isinstance(code, (int, float)) else str(code).strip()
        if isinstance(hire, datetime.datetime):
            code_to_hire[key] = hire.date()
        code_to_status[key] = status
        code_to_lastday[key] = as_date(last_day)
        code_to_resignnotif[key] = as_date(resign_notif)
        if position:
            code_to_position[key] = str(position).strip()
        if r[24]:
            code_to_email[key] = str(r[24]).strip().lower()
        if name:
            name_to_code[norm_name(name)] = key
        if position and 'brand manager' in str(position).lower():
            bm_roster.append((key, name, str(position).strip(), str(line).strip() if line else None))
        code_to_orgchart[key] = dict(
            bu=str(r[9]).strip() if r[9] else None,
            mgr1_code=r[12], mgr1_name=r[13],
            mgr2_code=r[14], mgr2_name=r[15],
            mgr3_code=r[16], mgr3_name=r[17],
        )
    code_to_probation_passed = {c: probation_passed_date(h) for c, h in code_to_hire.items()}

    def _usable_mgr_name(name):
        if not name:
            return False
        return 'vacant' not in str(name).lower()

    def resolve_direct_manager(code):
        """Escalates Direct -> 2nd -> 3rd Manager (Database Shortcut.xlsx),
        skipping any slot on file as literally "Vacant" (an unfilled
        intermediate role) the same way the org chart itself would."""
        entry = code_to_orgchart.get(code)
        if not entry:
            return None, None
        for name_key, code_key in (('mgr1_name', 'mgr1_code'), ('mgr2_name', 'mgr2_code'), ('mgr3_name', 'mgr3_code')):
            nm = entry.get(name_key)
            if _usable_mgr_name(nm):
                mgr_code = entry.get(code_key)
                mgr_code = str(int(mgr_code)) if isinstance(mgr_code, (int, float)) else (str(mgr_code).strip() if mgr_code else None)
                return str(nm).strip(), mgr_code
        return None, None

    def is_active_for_period(code, period_end):
        last_day = code_to_lastday.get(code)
        if last_day is not None:
            if last_day <= period_end:
                return False, f"Last Day of Work {last_day.isoformat()} on/before period end"
            return True, None
        status = code_to_status.get(code)
        if status != 'Active':
            return False, f"DB status = {status!r} (no Last Day of Work on file)"
        return True, None

    def is_probation_passed_for_period(code, period_start):
        pp = code_to_probation_passed.get(code)
        if pp is None:
            return None, None
        return (pp <= period_start), pp

    # -----------------------------------------------------------------
    # 1. Coverage + Right Frequency per rep, restricted to EVAL_PERIOD_NAME.
    # -----------------------------------------------------------------
    log('reading Coverage cache ...')
    records = load_b64('records.data.js')
    dash = load_b64('dashboard.data.js')
    dims = dash['dimensions']

    F = dict(period=0, team=1, businessUnit=2, nsm=3, areaManager=4, manager=5,
              employee=6, specialty=7, klass=8, status=9, experience=10, type=11,
              coveredDoctor=12, rightFreq=13, visits=14, isActive=15, actualPlanX1000=16,
              plansCount=17, title=18, customerName=19, profile=20, frequency=21,
              lastVisitDate=22, area=23)

    titles = dims['titles']
    types = dims['types']
    statuses = dims['statuses']
    teams = dims['teams']
    employeeNames = dims['employeeNames']
    employeeCodes = dims['employeeCodes']

    titleIdx = titles.index('Medical Representative')
    salesRepTitleIdx = titles.index('Sales Representative')
    statusIdx = statuses.index('Active')
    standardTypeIdx = {types.index(t) for t in ['Contract', 'Doctor', 'Hospital'] if t in types}
    pharmacyTypeIdx = {types.index(t) for t in ['Pharmacy'] if t in types}

    team_checks = []
    for t in teams:
        canon = normalize_line(t)
        bu = line_to_bu(t)
        is_chc_sales = (bu == 'CHC' and canon == 'CHC_SALES')
        team_checks.append(dict(bu=bu, canon=canon, is_chc_sales=is_chc_sales,
                                 titleIdx=(salesRepTitleIdx if is_chc_sales else titleIdx),
                                 typeSet=(pharmacyTypeIdx if is_chc_sales else standardTypeIdx)))

    rep_cov = defaultdict(lambda: dict(coveredSum=0, rightFreqSum=0, rowCount=0))
    emp_to_manager, emp_to_areaManager, emp_to_nsm = {}, {}, {}
    for row in records['rows']:
        empIdx = row[F['employee']]
        emp_to_manager[empIdx] = dims['managers'][row[F['manager']]]
        emp_to_areaManager[empIdx] = dims['areaManagers'][row[F['areaManager']]]
        emp_to_nsm[empIdx] = dims['nsms'][row[F['nsm']]]

        if dims['periods'][row[F['period']]] != EVAL_PERIOD_NAME:
            continue
        if row[F['status']] != statusIdx:
            continue
        check = team_checks[row[F['team']]]
        if not check['bu']:
            continue
        if row[F['title']] != check['titleIdx']:
            continue
        if row[F['type']] not in check['typeSet']:
            continue
        if not row[F['isActive']]:
            continue
        a = rep_cov[empIdx]
        a['coveredSum'] += row[F['coveredDoctor']] or 0
        a['rightFreqSum'] += row[F['rightFreq']] or 0
        a['rowCount'] += 1
        a['team'] = teams[row[F['team']]]

    reps_coverage = {}
    for empIdx, a in rep_cov.items():
        if a['rowCount'] == 0:
            continue
        name = employeeNames[empIdx]
        code = employeeCodes[empIdx]
        reps_coverage[norm_name(name)] = dict(
            empIdx=empIdx, name=name, code=code, team=a['team'],
            coveragePct=(a['coveredSum'] / a['rowCount']) * 100,
            rightFreqPct=(a['rightFreqSum'] / a['rowCount']) * 100,
            rowCount=a['rowCount'],
        )

    # -----------------------------------------------------------------
    # 2. Sales Achievement, restricted to EVAL_MONTH_STR.
    # -----------------------------------------------------------------
    log('reading Sales cache ...')
    sales_cache = load_b64('sales.data.js')
    MONTH, LINE, BRAND, PROD, REP, MASK, VAL, TGT_VAL = 0, 1, 2, 3, 4, 17, 19, 21
    linesLk = sales_cache['lookups']['lines']
    repsLk = sales_cache['lookups']['reps']
    monthsLk = sales_cache['lookups']['months']
    # Sales cache's own per-rep "position" is a specific territory/specialty
    # label (e.g. "NEUROSCIENCE QALUBIA") -- far more useful on a leaderboard
    # than Database Shortcut's generic Position column (almost always just
    # "Medical Representative"). Falls back to code_to_position if a rep
    # has no Sales cache row this month.
    name_to_salesposition = {}
    rep_positions_lk = sales_cache['lookups'].get('rep_positions', [])
    for i, rname in enumerate(repsLk):
        if i < len(rep_positions_lk) and rep_positions_lk[i]:
            name_to_salesposition[norm_name(rname)] = rep_positions_lk[i]
    scenarioCoverage = sales_cache['meta'].get('scenarioCoverage', {})
    schemaVersion = sales_cache['meta'].get('schemaVersion', 0)

    def resolve_scenario(raw, requested):
        canon = normalize_line(raw)
        cov = scenarioCoverage.get(canon) or scenarioCoverage.get(raw)
        if not cov:
            return requested
        if cov.get(requested):
            return requested
        other = 'working' if requested == 'official' else 'official'
        return other if cov.get(other) else requested

    want_official_by_line = [resolve_scenario(l, 'official') == 'official' for l in linesLk]

    def include_target_row(mask, want_official):
        if (mask & 16) == 0:
            return True
        if schemaVersion < 3:
            return True
        return ((mask & 32) > 0) == want_official

    rep_sales_month = defaultdict(lambda: dict(val=0, tgtVal=0))
    # National, line-level val/target totals for the eval month -- used to
    # auto-calculate each Brand Manager's National Sales Achievement from
    # the specific line/brand they're responsible for (see Step 7).
    line_sales_month = defaultdict(lambda: dict(val=0, tgtVal=0))
    for r in sales_cache['rows']:
        if (r[MASK] & 2) > 0:
            continue
        month_str = monthsLk[r[MONTH]] if r[MONTH] is not None and r[MONTH] < len(monthsLk) else None
        if month_str != EVAL_MONTH_STR:
            continue
        want_official = include_target_row(r[MASK], want_official_by_line[r[LINE]])
        la = line_sales_month[r[LINE]]
        la['val'] += r[VAL] or 0
        if want_official:
            la['tgtVal'] += r[TGT_VAL] or 0
        repName = repsLk[r[REP]] if r[REP] is not None and r[REP] < len(repsLk) else None
        if not repName:
            continue
        key = norm_name(repName)
        a = rep_sales_month[key]
        a['val'] += r[VAL] or 0
        if want_official:
            a['tgtVal'] += r[TGT_VAL] or 0

    def line_sales_achievement(sales_line_name):
        """National actual/target achievement fraction for a Sales-cache line
        name this month, or None if the line is unknown or has no target."""
        if sales_line_name not in linesLk:
            return None
        idx = linesLk.index(sales_line_name)
        la = line_sales_month.get(idx)
        if not la or not la['tgtVal']:
            return None
        return la['val'] / la['tgtVal']

    # -----------------------------------------------------------------
    # 3. Point curves.
    # -----------------------------------------------------------------
    log('reading scaling scores.xlsx ...')
    wb_curves = openpyxl.load_workbook(SCALING_PATH, data_only=True)
    msr_ws = wb_curves['medical rep points scheme']
    sr_ws = wb_curves['SALES REP']
    msr_sales_curve = extract_curve(msr_ws, 1, 2)
    msr_rf_curve = extract_curve(msr_ws, 4, 5)
    msr_cov_curve = extract_curve(msr_ws, 10, 11)
    sr_sales_curve = extract_curve(sr_ws, 1, 2)
    sr_cov_curve_raw = extract_curve(sr_ws, 10, 11)

    # DM/DSM, ASM/NSM, and Brand Manager curves -- found in this same
    # workbook (sheets not read before 2026-08-15: 'dsm points scheme',
    # 'ASM&NSM SCHEME', 'Brand Managers'). Each converts a raw achievement
    # fraction (or, for Region, a raw count) into points already scaled to
    # that KPI's weight -- exactly the same "extract_curve + interp"
    # pattern as the Medical Rep / CHC Sales Rep curves above. Team Avg
    # curves in these sheets are confirmed pure-linear (x * max points),
    # so the existing team_avg_pts = team_avg/100*weight calc already
    # matches them exactly; only the non-Team-Avg KPI columns need curves.
    dsm_ws = wb_curves['dsm points scheme']
    asmnsm_ws = wb_curves['ASM&NSM SCHEME']
    bm_ws = wb_curves['Brand Managers']
    dm_fielddays_curve = extract_curve(dsm_ws, 4, 5)     # Field Working Days, domain 0.6-1.0 -> 1-10 pts
    dm_dvcoverage_curve = extract_curve(dsm_ws, 7, 8)    # Double Visit Coverage, domain 0.8-1.0 -> 1-10 pts
    dm_callsperdv_curve = extract_curve(dsm_ws, 10, 11)  # Avg Calls per DV Day, domain 0.9-1.0 -> 6.1-10 pts
    asmnsm_fielddays_curve = extract_curve(asmnsm_ws, 4, 5)  # Field Working Days, domain 0.6-1.0 -> 1-20 pts
    bm_ach_curve = extract_curve(bm_ws, 1, 2)            # National Sales Ach%, domain 0.7-1.3 -> 1-50 pts
    bm_region_curve = extract_curve(bm_ws, 4, 5)         # Regions Covered (count), domain 1-5 -> 4-20 pts
    bm_tactical_curve = extract_curve(bm_ws, 8, 9)       # Tactical Plan Execution, domain 0.7-1.0 -> 21-30 pts

    # Brand Manager's Database Shortcut "Line" -> Sales cache "lines" lookup.
    # The two sheets name lines slightly differently (e.g. DB says "CNS",
    # Sales cache calls the same line "NEUROSCIENCE"); built by matching
    # every line value seen on the 11 active BM roster rows against the
    # Sales cache's own line list (see investigation 2026-08-15).
    BM_LINE_TO_SALES_LINE = {
        'PEDIA/GYN': 'PEDIA', 'ORTHO II': 'ORTHO-II', 'DIABETES I': 'DIAB-I',
        'CVM II': 'CVM-II', 'DIABETES II': 'DIAB-II', 'GIT II': 'GIT-II',
        'GIT I': 'GIT-I', 'GIT III': 'GIT-III', 'DIABETES III': 'DIAB-III',
        'DERMA': 'Derma', 'CNS': 'NEUROSCIENCE',
    }

    # -----------------------------------------------------------------
    # 4. Score Medical Rep + CHC Sales Rep, with per-period probation +
    #    active/resignation gating.
    # -----------------------------------------------------------------
    log('scoring Medical Rep / CHC Sales Rep ...')
    results = []
    excluded = []
    empidx_to_result = {}

    for key, cov in reps_coverage.items():
        code = cov['code']
        canon_line = normalize_line(cov['team'])
        bu = line_to_bu(cov['team'])

        active_ok, inactive_reason = is_active_for_period(code, EVAL_PERIOD_END)
        if not active_ok:
            last_day = code_to_lastday.get(code)
            notif = code_to_resignnotif.get(code)
            excluded.append(dict(code=code, name=cov['name'], line=canon_line, bu=bu, reason='not-active-resigned',
                                  detail=inactive_reason,
                                  lastDay=last_day.isoformat() if last_day else None,
                                  resignationNotif=notif.isoformat() if notif else None))
            continue

        prob_ok, pp = is_probation_passed_for_period(code, EVAL_PERIOD_START)
        if prob_ok is False:
            excluded.append(dict(code=code, name=cov['name'], line=canon_line, bu=bu, reason='probation-not-passed',
                                  detail=f"passes {pp.isoformat()}, ranking period starts {EVAL_PERIOD_START.isoformat()}",
                                  lastDay=None, resignationNotif=None))
            continue

        sales = rep_sales_month.get(key)
        is_sales_rep = (bu == 'CHC' and canon_line == 'CHC_SALES')
        ach_pct = (sales['val'] / sales['tgtVal']) if sales and sales['tgtVal'] > 0 else None

        if is_sales_rep:
            sales_pts = interp(sr_sales_curve, ach_pct) if ach_pct is not None else None
            cov_pts_raw = interp(sr_cov_curve_raw, cov['coveragePct'] / 100)
            cov_pts = cov_pts_raw * 4 if cov_pts_raw is not None else None
            rf_pts = None
            total = (sales_pts or 0) + (cov_pts or 0)
            role = 'Sales Rep (CHC)'
        else:
            sales_pts = interp(msr_sales_curve, ach_pct) if ach_pct is not None else None
            cov_pts = interp(msr_cov_curve, cov['coveragePct'] / 100)
            rf_pts = interp(msr_rf_curve, cov['rightFreqPct'] / 100)
            total = (sales_pts or 0) + (cov_pts or 0) + (rf_pts or 0)
            role = 'Medical Rep'

        last_day = code_to_lastday.get(code)
        is_departing_soon = last_day is not None and last_day > EVAL_PERIOD_END

        rec = dict(
            name=cov['name'], code=code, team=cov['team'], canonLine=canon_line, bu=bu, role=role,
            position=name_to_salesposition.get(key) or code_to_position.get(code),
            hireDate=code_to_hire.get(code).isoformat() if code_to_hire.get(code) else None,
            probationPassed=pp.isoformat() if pp else None,
            achPct=ach_pct, coveragePct=cov['coveragePct'], rightFreqPct=cov['rightFreqPct'],
            salesPts=sales_pts, covPts=cov_pts, rfPts=rf_pts, totalPts=total,
            isDepartingSoon=is_departing_soon,
            departingLastDay=last_day.isoformat() if is_departing_soon else None,
            # Raw sales value/target (not the ach_pct ratio) -- carried up
            # through DM/DSM -> ASM/NSM so each manager tier's own "team
            # floor" can be computed as SUM(val)/SUM(tgtVal) across every
            # rep beneath them, never as an average of already-divided
            # percentages. Per Ahmed 2026-08-16 ("DM DSM HAS REWARD ALSO,
            # MAKE SAME FLOOR LOGIC FOR ASM NSM"). 0/0 when a rep has no
            # matched sales row this month -- contributes nothing to
            # either side of the team sum, exactly as it should.
            salesVal=(sales['val'] if sales else 0.0) or 0.0,
            salesTgt=(sales['tgtVal'] if sales else 0.0) or 0.0,
        )
        results.append(rec)
        empidx_to_result[cov['empIdx']] = rec

    # "Departing soon" -- ranked reps who already have a resignation on
    # file with a Last Day after this period (still correctly included,
    # but useful to see who rolls off and when). Also flagged inline on
    # each rep's own record above (isDepartingSoon) so it can show as a
    # badge directly on the ranked row, per Ahmed 2026-08-15.
    departing_soon = []
    for rec in results:
        if not rec['isDepartingSoon']:
            continue
        code = rec['code']
        departing_soon.append(dict(code=code, name=rec['name'], line=rec['canonLine'], bu=rec['bu'],
                                    resignationNotif=code_to_resignnotif.get(code).isoformat() if code_to_resignnotif.get(code) else None,
                                    lastDay=rec['departingLastDay']))
    departing_soon.sort(key=lambda d: d['lastDay'])

    log(f'  Medical Rep + Sales Rep: {len(results)} ranked, {len(excluded)} excluded, '
        f'{len(departing_soon)} departing soon')

    # -----------------------------------------------------------------
    # 5. Missing-KPI template: auto-created once (never overwritten, so
    #    Ahmed's fills always survive a re-run), read back in every run.
    # -----------------------------------------------------------------
    def load_kpi_template():
        if not os.path.exists(TEMPLATE_PATH):
            return {}
        wb_t = openpyxl.load_workbook(TEMPLATE_PATH, data_only=True)
        out = {}
        for sheet_name, col_map in TEMPLATE_SHEETS.items():
            if sheet_name not in wb_t.sheetnames:
                continue
            ws_t = wb_t[sheet_name]
            header = [c.value for c in next(ws_t.iter_rows(min_row=1, max_row=1))]
            if 'Code' not in header:
                continue
            code_col = header.index('Code')
            sheet_data = {}
            for row in ws_t.iter_rows(min_row=2, values_only=True):
                code = row[code_col]
                if code is None:
                    continue
                code = str(int(code)) if isinstance(code, (int, float)) else str(code).strip()
                vals = {}
                for key, col_label, _weight in col_map:
                    if col_label in header:
                        v = row[header.index(col_label)]
                        if v is not None and str(v).strip() != '':
                            try:
                                vals[key] = float(v)
                            except (TypeError, ValueError):
                                pass
                if vals:
                    sheet_data[code] = vals
            out[sheet_name] = sheet_data
        return out

    def write_kpi_template(dm_roster, asm_roster, nsm_roster, bm_roster_active):
        if os.path.exists(TEMPLATE_PATH):
            log(f'KPI template already exists at {TEMPLATE_PATH} -- leaving it untouched '
                f'(delete it manually to force a fresh blank template).')
            return
        wb_t = openpyxl.Workbook()
        ws0 = wb_t.active
        ws0.title = 'Instructions'
        ws0.append(['Zeta Sprint 2026 -- Missing KPI Template'])
        ws0.append([])
        ws0.append(['Fill in the columns on each tab below, save this file in place, and tell'])
        ws0.append(['Claude -- the next cache rebuild (etl/build_sprint_cache.py) picks it up'])
        ws0.append(['automatically by employee Code. Leave a cell blank if you don\'t have that'])
        ws0.append(['number yet; it will keep showing as "pending" until filled.'])
        ws0.append([])
        ws0.append(['IMPORTANT: each column asks for the REAL ACHIEVEMENT number, not points.'])
        ws0.append(['Enter a decimal fraction, e.g. "Field Working Days -- Actual % Achieved'])
        ws0.append(['(e.g. 0.85 for 85%)" means enter 0.85 for 85% -- not 85. The one exception'])
        ws0.append(['is "Regions Covered -- Actual Count (1-5)", which is a whole number.'])
        ws0.append(['build_sprint_cache.py converts your number into points itself, using the'])
        ws0.append(['real scoring curves from zeta sprint/scaling scores.xlsx (sheets "dsm points'])
        ws0.append(['scheme", "ASM&NSM SCHEME", "Brand Managers") -- the same curves that already'])
        ws0.append(['score Medical Rep / CHC Sales Rep Sales Achievement, Right Frequency, and'])
        ws0.append(['Coverage. You do not need to do any scaling math yourself.'])
        ws0.append([])
        ws0.append(['Code and Name are pre-filled from Database Shortcut.xlsx for everyone currently'])
        ws0.append(['eligible this ranking period -- do not need to add/remove rows for a normal month.'])
        ws0.append([])
        ws0.append(['Brand_Manager: National Sales is NOT in this template -- it is auto-calculated'])
        ws0.append(['from the Sales cache using the specific Line/brand each Brand Manager is'])
        ws0.append(['responsible for (per Database Shortcut.xlsx), so there is nothing to fill in'])
        ws0.append(['for it. Only Regions Covered and Tactical Plan Execution need your input.'])

        def add_sheet(name, rows, extra_cols):
            ws = wb_t.create_sheet(name)
            header = ['Code', 'Name'] + [c for _k, c, _w in extra_cols]
            ws.append(header)
            for row in rows:
                ws.append(list(row) + [None] * len(extra_cols))
            widths = [10, 34] + [28] * len(extra_cols)
            for i, w in enumerate(widths, start=1):
                ws.column_dimensions[chr(64 + i) if i <= 26 else 'A'].width = w

        add_sheet('DM_DSM', [(r['code'], r['name']) for r in dm_roster], TEMPLATE_SHEETS['DM_DSM'])
        add_sheet('ASM', [(r['code'], r['name']) for r in asm_roster], TEMPLATE_SHEETS['ASM'])
        add_sheet('NSM', [(r['code'], r['name']) for r in nsm_roster], TEMPLATE_SHEETS['NSM'])
        add_sheet('Brand_Manager', [(c, n) for c, n, _p, _l in bm_roster_active], TEMPLATE_SHEETS['Brand_Manager'])

        os.makedirs(os.path.dirname(TEMPLATE_PATH), exist_ok=True)
        wb_t.save(TEMPLATE_PATH)
        log(f'wrote NEW KPI template: {TEMPLATE_PATH}')

    kpi_template = load_kpi_template()

    def normalize_raw(key, val):
        """Defensive guard against a common fill-in mistake: entering a
        whole-number percent (e.g. 85) instead of the requested decimal
        fraction (0.85). Every raw-achievement KPI here has a legitimate
        range well under 3 (curves top out around 1.3 for over-achieved
        Sales); 'regionCount' is the one genuinely whole-number field
        (1-5) and must NOT be rescaled."""
        if key == 'regionCount':
            return val
        return val / 100.0 if val > 3 else val

    # -----------------------------------------------------------------
    # 6. DM/DSM, ASM, NSM -- Team Avg rollup (this month's already-scored
    #    reps) + whatever KPI slots Ahmed has filled into the template,
    #    converted from raw achievement to points via the real curves
    #    from scaling scores.xlsx.
    # -----------------------------------------------------------------
    log('rolling up DM/DSM, ASM, NSM ...')

    def score_hierarchy_tier(name_list, team_pool, team_avg_weight, kpi_slots, tier_label, template_sheet, curves,
                              member_noun='rep'):
        # team_pool: dict of mgr_name -> list of member dicts, each already
        # normalized to {name, code, line, bu, role, totalPts}. For DM/DSM
        # the pool is individual reps; for ASM/NSM the pool is DM/DSMs
        # (see call sites below) -- per Ahmed 2026-08-15 ("ASM and NSM
        # average points for their DSM"), ASM/NSM Team Avg now rolls up
        # from each DM/DSM's own total points, not straight from reps,
        # matching the real org chain Rep -> DM/DSM -> ASM/NSM.
        # Only members who themselves passed the active + probation gates
        # (for reps) or were successfully scored (for DM/DSMs) ever appear
        # in a pool, so this Team Avg is already restricted to eligible
        # members only -- confirmed 2026-08-15 per Ahmed's question.
        provided_by_code = kpi_template.get(template_sheet, {})
        out = []
        excl = []
        for mgr_name in name_list:
            if not mgr_name or mgr_name.strip() == '' or mgr_name.upper().startswith('VACANT'):
                continue
            code = name_to_code.get(norm_name(mgr_name))
            team_members = team_pool.get(mgr_name, [])
            line_counts = Counter(m['line'] for m in team_members if m.get('line'))
            bu_counts = Counter(m['bu'] for m in team_members if m.get('bu'))
            primary_line = line_counts.most_common(1)[0][0] if line_counts else None
            primary_bu = bu_counts.most_common(1)[0][0] if bu_counts else None
            team_members_out = [dict(name=m['name'], code=m['code'], line=m['line'], bu=m['bu'],
                                      role=m['role'], totalPts=m['totalPts'])
                                 for m in team_members if m['totalPts'] is not None]
            team_members_out.sort(key=lambda m: m['totalPts'], reverse=True)

            if code is None:
                excl.append(dict(name=mgr_name, code=None, reason='no-database-match',
                                  detail='Name not found in Database Shortcut.xlsx -- cannot verify probation/active status',
                                  line=primary_line, bu=primary_bu))
                continue

            active_ok, inactive_reason = is_active_for_period(code, EVAL_PERIOD_END)
            if not active_ok:
                excl.append(dict(name=mgr_name, code=code, reason='not-active-resigned', detail=inactive_reason,
                                  line=primary_line, bu=primary_bu))
                continue
            prob_ok, pp = is_probation_passed_for_period(code, EVAL_PERIOD_START)
            if prob_ok is False:
                excl.append(dict(name=mgr_name, code=code, reason='probation-not-passed',
                                  detail=f"passes {pp.isoformat()}, ranking period starts {EVAL_PERIOD_START.isoformat()}",
                                  line=primary_line, bu=primary_bu))
                continue

            team_scores = [m['totalPts'] for m in team_members if m['totalPts'] is not None]
            team_avg = (sum(team_scores) / len(team_scores)) if team_scores else None
            team_avg_pts = (team_avg / 100 * team_avg_weight) if team_avg is not None else None

            # Team Sales Achievement % -- SUM(val)/SUM(tgtVal) across every
            # eligible member, divided exactly once (never an average of
            # per-member percentages). For DM/DSM, members are individual
            # reps carrying their own raw salesVal/salesTgt. For ASM/NSM,
            # members are DM/DSMs carrying THEIR team's already-summed
            # salesVal/salesTgt (set below in the ASM/NSM pool build) --
            # so this same sum-once logic rolls all the way up from the
            # rep level with no percentage ever averaged at any tier.
            # Gates the 70% floor for WINNER/RUNNER-UP + BU Leader badges
            # and the cash payout -- per Ahmed 2026-08-16.
            team_sales_val = sum(m.get('salesVal', 0) or 0 for m in team_members)
            team_sales_tgt = sum(m.get('salesTgt', 0) or 0 for m in team_members)
            team_sales_ach_pct = (team_sales_val / team_sales_tgt) if team_sales_tgt > 0 else None

            provided = provided_by_code.get(code, {})
            kpis = []
            extra_pts_sum = 0.0
            any_pending = team_avg_pts is None
            for key, label, weight in kpi_slots:
                raw = provided.get(key)
                pts = None
                if raw is not None:
                    raw = normalize_raw(key, raw)
                    pts = max(0.0, min(weight, interp(curves[key], raw)))
                    extra_pts_sum += pts
                else:
                    any_pending = True
                kpis.append(dict(key=key, label=label, weight=weight, pts=pts, raw=raw))

            total_pts = None if team_avg_pts is None else (team_avg_pts + extra_pts_sum)

            out.append(dict(
                name=mgr_name, code=code, tier=tier_label, line=primary_line, bu=primary_bu,
                hireDate=code_to_hire.get(code).isoformat() if code_to_hire.get(code) else None,
                probationPassed=pp.isoformat() if pp else None,
                teamSize=len(team_scores),
                memberNoun=member_noun,
                teamMembers=team_members_out,
                teamAvgRaw=team_avg,
                teamAvgPts=team_avg_pts,
                teamAvgWeight=team_avg_weight,
                kpis=kpis,
                totalPts=total_pts,
                totalMaxPts=100,
                isPartial=any_pending,
                teamSalesVal=team_sales_val,
                teamSalesTgt=team_sales_tgt,
                teamSalesAchPct=team_sales_ach_pct,
            ))
        return out, excl

    # DM/DSM: Team Avg 70 + Field Working Days 10 + DV Coverage 10 +
    # Calls per DV 10 -- confirmed by Ahmed 2026-08-15 ("dm dsm is right
    # for these parameters as in scaling and in ppt") against the real
    # 'dsm points scheme' curves. Team Avg pool = individual reps under
    # each DM/DSM.
    dm_curves = {'fieldDays': dm_fielddays_curve, 'dvCoverage': dm_dvcoverage_curve, 'callsPerDv': dm_callsperdv_curve}
    dm_team_pool = defaultdict(list)
    for empIdx, rec in empidx_to_result.items():
        mgr_name = emp_to_manager.get(empIdx)
        if mgr_name:
            dm_team_pool[mgr_name].append(dict(name=rec['name'], code=rec['code'], line=rec['canonLine'],
                                                bu=rec['bu'], role=rec['role'], totalPts=rec['totalPts'],
                                                salesVal=rec['salesVal'], salesTgt=rec['salesTgt']))
    dm_results, dm_excluded = score_hierarchy_tier(dims['managers'], dm_team_pool, 70,
                                                     TEMPLATE_SHEETS['DM_DSM'], 'DM/DSM', 'DM_DSM', dm_curves,
                                                     member_noun='rep')

    # ASM/NSM: (Sub-)Team Avg 80 + Field Days 20, per 'ASM&NSM SCHEME' --
    # weights unchanged, but per Ahmed 2026-08-15 ("ASM and NSM average
    # points for their DSM") the Team Avg pool is now each ASM/NSM's own
    # DM/DSMs (their totalPts, out of 100), not individual reps directly
    # -- matches the real org chain Rep -> DM/DSM -> ASM/NSM. Each DM/DSM's
    # ASM/NSM is derived by majority vote over their own (already-eligible)
    # reps' direct ASM/NSM assignment in the Coverage cache -- the same
    # majority-vote technique already used for each manager's Line/BU.
    asmnsm_curves = {'fieldDays': asmnsm_fielddays_curve}
    dm_asm_votes = defaultdict(Counter)
    dm_nsm_votes = defaultdict(Counter)
    for empIdx, rec in empidx_to_result.items():
        dm_name = emp_to_manager.get(empIdx)
        if not dm_name:
            continue
        asm_name = emp_to_areaManager.get(empIdx)
        nsm_name = emp_to_nsm.get(empIdx)
        if asm_name:
            dm_asm_votes[dm_name][asm_name] += 1
        if nsm_name:
            dm_nsm_votes[dm_name][nsm_name] += 1
    dm_name_to_asm = {dm: votes.most_common(1)[0][0] for dm, votes in dm_asm_votes.items()}
    dm_name_to_nsm = {dm: votes.most_common(1)[0][0] for dm, votes in dm_nsm_votes.items()}

    asm_team_pool = defaultdict(list)
    nsm_team_pool = defaultdict(list)
    for dm in dm_results:
        if dm['totalPts'] is None:
            continue
        member = dict(name=dm['name'], code=dm['code'], line=dm['line'], bu=dm['bu'],
                      role='DM/DSM', totalPts=dm['totalPts'],
                      salesVal=dm['teamSalesVal'], salesTgt=dm['teamSalesTgt'])
        asm_name = dm_name_to_asm.get(dm['name'])
        if asm_name:
            asm_team_pool[asm_name].append(dict(member))
        nsm_name = dm_name_to_nsm.get(dm['name'])
        if nsm_name:
            nsm_team_pool[nsm_name].append(dict(member))

    asm_results, asm_excluded = score_hierarchy_tier(dims['areaManagers'], asm_team_pool, 80,
                                                       TEMPLATE_SHEETS['ASM'], 'ASM', 'ASM', asmnsm_curves,
                                                       member_noun='DM/DSM')
    nsm_results, nsm_excluded = score_hierarchy_tier(dims['nsms'], nsm_team_pool, 80,
                                                       TEMPLATE_SHEETS['NSM'], 'NSM', 'NSM', asmnsm_curves,
                                                       member_noun='DM/DSM')

    log(f'  DM/DSM: {len(dm_results)} scored, {len(dm_excluded)} excluded')
    log(f'  ASM:    {len(asm_results)} scored, {len(asm_excluded)} excluded')
    log(f'  NSM:    {len(nsm_results)} scored, {len(nsm_excluded)} excluded')

    # -----------------------------------------------------------------
    # 7. Brand Manager -- roster from Database Shortcut (owns products,
    #    not a rep hierarchy, so no Team Avg rollup is possible).
    #    National Sales (Ach%) is auto-calculated from the Sales cache,
    #    aggregated nationally over the specific Line/brand this Brand
    #    Manager is responsible for, per Ahmed 2026-08-15 ("calculate bm
    #    achievement according to brands they are responsible for").
    #    Regions Covered + Tactical Plan Execution still come from the
    #    template (raw achievement -> points via the Brand Managers curve).
    # -----------------------------------------------------------------
    log('scoring Brand Manager ...')
    bm_template = kpi_template.get('Brand_Manager', {})
    bm_results, bm_excluded = [], []
    for code, name, position, line in bm_roster:
        active_ok, inactive_reason = is_active_for_period(code, EVAL_PERIOD_END)
        if not active_ok:
            bm_excluded.append(dict(name=name, code=code, reason='not-active-resigned', detail=inactive_reason))
            continue
        prob_ok, pp = is_probation_passed_for_period(code, EVAL_PERIOD_START)
        if prob_ok is False:
            bm_excluded.append(dict(name=name, code=code, reason='probation-not-passed',
                                     detail=f"passes {pp.isoformat()}, ranking period starts {EVAL_PERIOD_START.isoformat()}"))
            continue

        sales_line = BM_LINE_TO_SALES_LINE.get(line.strip().upper()) if line else None
        ach = line_sales_achievement(sales_line) if sales_line else None
        kpis = []
        total_pts = 0.0
        any_pending = False
        if ach is not None:
            pts = max(0.0, min(50.0, interp(bm_ach_curve, ach)))
            total_pts += pts
        else:
            pts = None
            any_pending = True
        kpis.append(dict(key='salesAch', label='National Sales -- Ach% (auto)', weight=50, pts=pts, raw=ach,
                          source='auto', line=line, salesLine=sales_line))

        provided = bm_template.get(code, {})
        bm_curves = {'regionCount': bm_region_curve, 'tacticalPlan': bm_tactical_curve}
        for key, label, weight in TEMPLATE_SHEETS['Brand_Manager']:
            raw = provided.get(key)
            if raw is not None:
                raw = normalize_raw(key, raw)
                pts = max(0.0, min(weight, interp(bm_curves[key], raw)))
                total_pts += pts
            else:
                pts = None
                any_pending = True
            kpis.append(dict(key=key, label=label, weight=weight, pts=pts, raw=raw, source='template'))

        bm_results.append(dict(
            name=name, code=code, tier='Brand Manager', position=position, line=line,
            # bu reuses the same sales_line -> canonical-BU lookup already
            # computed above for National Sales -- Brand Managers' own raw
            # Database Shortcut Line strings ("DIABETES I", etc.) don't match
            # the canonical Coverage-cache line vocabulary directly, but the
            # BM_LINE_TO_SALES_LINE-translated sales_line does. Powers BU
            # scoping for a Line/BU-restricted dashboard user, per Ahmed
            # 2026-08-15.
            bu=line_to_bu(sales_line) if sales_line else None,
            hireDate=code_to_hire.get(code).isoformat() if code_to_hire.get(code) else None,
            probationPassed=pp.isoformat() if pp else None,
            kpis=kpis,
            totalPts=total_pts,
            totalMaxPts=100,
            isPartial=any_pending,
        ))

    log(f'  Brand Manager: {len(bm_results)} scored, {len(bm_excluded)} excluded')

    write_kpi_template(dm_results, asm_results, nsm_results,
                        [(c, n, p, l) for c, n, p, l in bm_roster if any(r['code'] == c for r in bm_results)])

    # -----------------------------------------------------------------
    # 7b. Direct Manager + Direct Manager BU, every tier -- sourced from
    #     Database Shortcut.xlsx's own org-chart columns (resolve_direct_
    #     manager above), which cover every employee code including ASM/
    #     NSM -- per Ahmed 2026-08-15 ("asm and nsm name of manager and bu
    #     not present"). BU for the resolved manager prefers this cache's
    #     own canonical bu label (keeps it consistent with the rest of the
    #     dashboard, e.g. "DIAB" not Database Shortcut's raw "Diabetes"),
    #     falling back to Database Shortcut's raw Business Unit column only
    #     when the manager isn't scored anywhere in this cache (e.g. above
    #     NSM level, where there's no sprint tier to canonicalize against).
    # -----------------------------------------------------------------
    code_to_bu = {}
    for rec in results + dm_results + asm_results + nsm_results:
        if rec.get('bu'):
            code_to_bu[rec['code']] = rec['bu']

    def attach_manager_fields(rec_list):
        for rec in rec_list:
            mgr_name, mgr_code = resolve_direct_manager(rec['code'])
            mgr_bu = None
            if mgr_name:
                mgr_bu = code_to_bu.get(mgr_code)
                if not mgr_bu and mgr_code:
                    entry = code_to_orgchart.get(mgr_code)
                    if entry and entry.get('bu'):
                        mgr_bu = entry['bu']
            rec['directManager'] = mgr_name
            rec['directManagerBu'] = mgr_bu
            # Business Email -- lets the frontend match a logged-in dashboard
            # user back to their own record for the "My Performance" card.
            rec['email'] = code_to_email.get(rec['code'])

    attach_manager_fields(results)
    attach_manager_fields(dm_results)
    attach_manager_fields(asm_results)
    attach_manager_fields(nsm_results)
    attach_manager_fields(bm_results)
    log('attached Direct Manager / Direct Manager BU / email to every tier (Database Shortcut.xlsx)')

    # -----------------------------------------------------------------
    # 6. Assemble + write cache.
    # -----------------------------------------------------------------
    cache = {
        'meta': {
            'schemaVersion': SCHEMA_VERSION,
            'generatedAt': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'evalPeriod': EVAL_PERIOD_NAME,
            'periodStart': EVAL_PERIOD_START.isoformat(),
            'periodEnd': EVAL_PERIOD_END.isoformat(),
            'methodology': {
                'probationRule': 'Hire day 1-15 -> ref = 1st of same month; 16-31 -> ref = 1st of next month; '
                                  'probation passes 3 months after ref. Eligible for month M if pass-date <= 1st of M.',
                'activeRule': 'Excluded if Last Day of Work (Database Shortcut.xlsx) is on/before the period end; '
                              'Status field used only when no Last Day is on file.',
                'curveSheet': 'medical rep points scheme / SALES REP',
                'chcCoverageScaling': 'x4 on the raw 10pt Coverage curve to hit the deck\'s 40pt weight (flagged, unconfirmed with workbook author)',
                'confirmedWithAhmed': '2026-08-15',
            },
            'tiersBuilt': ['Medical Rep', 'Sales Rep (CHC)', 'DM/DSM', 'ASM', 'NSM', 'Brand Manager'],
            'tiersPending': [],
            'pendingDataNotes': 'DM/DSM, ASM, NSM and Brand Manager are all live, but each still has KPI slots '
                                 '(Field Working Days / DV Coverage / Calls-per-DV for DM-ASM-NSM; Sales '
                                 'Achievement / Regional Coverage / Tactical Plan Execution for Brand Manager) '
                                 'sourced from zeta sprint/Sprint_Missing_KPI_Template.xlsx, filled in by Ahmed. '
                                 'Any slot still blank there shows as "pending" here, never as zero.',
            'kpiTemplatePath': 'zeta sprint/Sprint_Missing_KPI_Template.xlsx',
            # Period-filter scaffolding, per Ahmed 2026-08-15: June is the
            # Sprint's Pilot Period; July/August are the first two real
            # months; Q3 = Jul+Aug+Sep cumulative; Oct/Nov individually;
            # Total Year = Q3+Oct+Nov+Dec cumulative. Only calendar months
            # actually present in the Coverage/Sales caches can be scored
            # -- as of this build that's only 'periodsBuilt' below. The
            # frontend uses this list to show every other period option as
            # "not yet available" rather than guessing at numbers with no
            # source data behind them. Re-run this script (with
            # EVAL_PERIOD_NAME advanced) once each new month's data lands
            # in the main dashboard cache; the multi-month cumulative
            # scoring engine that sums per-month results for Q3/Total Year
            # is a follow-up build once there's real July data to validate
            # it against -- not guessed at now.
            'periodsBuilt': [EVAL_PERIOD_NAME],
            'sourceDataPeriodsAvailable': dims['periods'],
        },
        'medicalRepSalesRep': {
            'ranked': results,
            'excluded': excluded,
            'departingSoon': departing_soon,
        },
        'dmDsm': {'ranked': dm_results, 'excluded': dm_excluded},
        'asm': {'ranked': asm_results, 'excluded': asm_excluded},
        'nsm': {'ranked': nsm_results, 'excluded': nsm_excluded},
        'brandManager': {'ranked': bm_results, 'excluded': bm_excluded},
    }

    json_str = json.dumps(cache, separators=(',', ':'), ensure_ascii=False, default=str)
    gz = gzip.compress(json_str.encode('utf-8'), compresslevel=9)
    b64 = base64.b64encode(gz).decode('ascii')

    os.makedirs(CACHE_DIR, exist_ok=True)

    tmp = OUT_JSON + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        f.write(json_str)
    os.replace(tmp, OUT_JSON)

    tmp = OUT_JS + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        f.write('window.SPRINT_CACHE = {b64Data:"' + b64 + '"};\n')
    os.replace(tmp, OUT_JS)

    log(f'wrote {os.path.basename(OUT_JSON)}  {os.path.getsize(OUT_JSON) // 1024:,} KB')
    log(f'wrote {os.path.basename(OUT_JS)}  {os.path.getsize(OUT_JS) // 1024:,} KB (gzip+base64)')

    # -----------------------------------------------------------------
    # 6b. Archive this month's full snapshot + update the history index,
    #     so past months stay viewable in the Period filter after a later
    #     month's re-run overwrites sprint.json/sprint.data.js -- per
    #     Ahmed 2026-08-15 ("if I need to make months update what should
    #     be"). Archived as a `window.X = {...}` script (like every other
    #     *.data.js cache here), NOT plain JSON fetched at runtime --
    #     this dashboard is opened over file:// in production, where
    #     Chrome blocks fetch()/XHR to local files but allows
    #     <script src="local/file.js">. index.json is Python-side
    #     bookkeeping only; index.js is what the browser reads.
    # -----------------------------------------------------------------
    os.makedirs(HISTORY_DIR, exist_ok=True)
    period_key = EVAL_PERIOD_START.strftime('%Y-%m')

    archive_js = os.path.join(HISTORY_DIR, f'sprint_{period_key}.data.js')
    tmp = archive_js + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        f.write('window.SPRINT_HISTORY_CACHE = {b64Data:"' + b64 + '"};\n')
    os.replace(tmp, archive_js)

    index_json_path = os.path.join(HISTORY_DIR, 'index.json')
    manifest = []
    if os.path.exists(index_json_path):
        try:
            with open(index_json_path, 'r', encoding='utf-8') as f:
                manifest = json.load(f)
        except Exception as e:
            log(f'WARNING: could not read existing history index ({e}); starting a fresh one')
            manifest = []
    manifest = [p for p in manifest if p.get('key') != period_key]
    manifest.append({
        'key': period_key,
        'name': EVAL_PERIOD_NAME,
        'file': f'cache/sprint_history/sprint_{period_key}.data.js',
        'generatedAt': datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    })
    manifest.sort(key=lambda p: p['key'])

    tmp = index_json_path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, indent=2)
    os.replace(tmp, index_json_path)

    tmp = HISTORY_INDEX_JS + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        f.write('window.SPRINT_HISTORY_INDEX = ' + json.dumps(manifest, separators=(',', ':')) + ';\n')
    os.replace(tmp, HISTORY_INDEX_JS)

    log(f'archived {period_key} ({EVAL_PERIOD_NAME}) -> {os.path.basename(archive_js)}; '
        f'history now covers: {", ".join(p["name"] for p in manifest)}')
    log(f'Sprint cache complete in {time.time() - t0:.1f}s')


if __name__ == '__main__':
    main()
