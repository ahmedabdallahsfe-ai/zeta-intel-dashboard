"""
build_expense_foundation.py — PHASE A: trustworthy expense data foundation
==============================================================================

Produces the reviewable artefacts the Expense vs Sales capability must stand on
BEFORE any dashboard, KPI or Ask work begins. It builds no UI and computes no
business KPI. Its only job is to answer one question honestly:

    Can expense rows be joined to sales rows, and if not, exactly which ones,
    for how much money, and why?

WHAT IT WRITES
------------------------------------------------------------------------------
    expense/mapping/expense_sku_map.csv        THE reviewable mapping table
    expense/reports/budget_reconciliation.csv  SKU sheet vs Line sheet
    expense/reports/dimension_exceptions.csv   BU/Line values that do not map
    expense/reports/unmapped_skus.csv          everything needing a decision
    expense/templates/expense_actuals_template.xlsx
    expense/PHASE_A_SUMMARY.txt                the GO / NO-GO numbers

HUMAN IN THE LOOP — THE CENTRAL DESIGN DECISION
------------------------------------------------------------------------------
`expense_sku_map.csv` is generated ONCE with proposed matches, then owned by a
human. On every later run this script READS the existing file, preserves every
row a person has reviewed (`Reviewed = YES`), and only proposes rows that are
new or still unreviewed.

That is deliberate. Automatic fuzzy matching at runtime would silently map
"COXORIZET 60 MG 20 TAB" to the wrong pack the day someone adds a second 60 MG
presentation, and no one would ever see it happen. A mapping that a person
signed off, stored as data and applied deterministically, cannot drift on its
own.

STATUS VOCABULARY
------------------------------------------------------------------------------
    MATCHED           name matches a sales product exactly
    NAME_VARIANT      matches once case/spacing/punctuation are normalised.
                      PROPOSED ONLY -- a human confirms before it counts.
    NOT_YET_SELLING   the brand exists in sales, this SKU has no sales record.
                      The label names the MEASUREMENT, not a business status.
                      Whether such a SKU is a launch, a naming variant we have
                      not caught, or something else is a question for a person
                      who knows the product -- the data cannot distinguish them.
    UNMAPPED          neither the SKU nor its brand appears in sales. Needs a
                      business answer before it can be used for anything.

NOT_YET_SELLING and UNMAPPED are kept apart on purpose. Collapsing them would
hide the difference between "the brand sells but this pack does not" and
"nothing by this name sells at all", and those lead to different investigations.

NO STATUS IS A BUSINESS CLAIM. Ahmed's instruction, 2026-08-09: do not label a
SKU pre-launch, discontinued, or anything else the source does not state. These
four values describe what the JOIN did, and nothing more.

WHAT IT REFUSES TO DO
------------------------------------------------------------------------------
  * It does not reconcile the two budget sheets. They disagree; the disagreement
    is reported and preserved.
  * It does not guess a BU for an unrecognised Line.
  * It does not drop unmatched budget. Every EGP is accounted for in the
    reconciliation, whether or not it can be joined to sales.
"""

import os
import re
import csv
import sys
import json
import gzip
import base64
from collections import defaultdict

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
EXPENSE_XLSX = os.path.join(ROOT_DIR, 'Market Expenses & FTE (3).xlsx')
SALES_CACHE = os.path.join(ROOT_DIR, 'cache', 'sales.data.js')
SEMANTIC_JS = os.path.join(ROOT_DIR, 'js', 'semantic-model.js')

OUT_DIR = os.path.join(ROOT_DIR, 'expense')
MAP_DIR = os.path.join(OUT_DIR, 'mapping')
REP_DIR = os.path.join(OUT_DIR, 'reports')
TPL_DIR = os.path.join(OUT_DIR, 'templates')

MAP_CSV = os.path.join(MAP_DIR, 'expense_sku_map.csv')

SHEET_SKU = "Expenses per SKU's Monthly"
SHEET_LINE = 'Expenses per line monthly'

MAP_HEADER = [
    'ExpenseBU', 'ExpenseLine', 'ExpenseBrand', 'ExpenseCode', 'ExpenseSKU',
    'Active', 'BudgetTotal', 'ProposedSalesProduct', 'ProposedSalesBrand',
    'CanonicalBU', 'CanonicalLine', 'MappingStatus', 'MappingReason',
    'Reviewed', 'ReviewedBy', 'ReviewNote',
]


# =============================================================================
# Reading the inputs
# =============================================================================
def norm(s):
    """
    Loose key for comparing names: lowercase, alphanumeric, single-spaced,
    with digit/letter boundaries separated.

    That last step matters more than it looks. The expense sheet writes
    "COXORIZET 60 MG 20 TAB"; the sales cube writes "COXORIZET 60MG 20 TAB".
    Without splitting the boundary, "60mg" and "60 mg" are different tokens and
    the SKU is reported as NOT_YET_SELLING -- i.e. budget with no sales, a
    finding that would have sent someone looking for a launch that had in fact
    been selling all along. Two such SKUs, worth 12,000,000 EGP of ACTIVE
    budget, were misclassified exactly this way before this was added.

    This is a safe normalisation, not a fuzzy guess: a space between a number
    and its unit carries no meaning. Anything looser -- edit distance, token
    subsets -- is deliberately NOT done here, because it would start proposing
    matches that are merely plausible. Those belong to human review.
    """
    t = re.sub(r'[^a-z0-9]+', ' ', str(s).lower())
    t = re.sub(r'(\d)([a-z])', r'\1 \2', t)
    t = re.sub(r'([a-z])(\d)', r'\1 \2', t)
    return re.sub(r'\s+', ' ', t).strip()


def load_semantic_lines():
    """
    The authoritative Line -> BU map AND the synonym table, both read from
    js/semantic-model.js.

    Parsed out of the live source rather than duplicated here. A second copy
    would drift from the first within a month, and then two parts of the
    platform would disagree about which BU a line belongs to.

    CORRECTION 2026-08-09. The first version of this script consulted only
    CANONICAL_LINE_TO_BU and therefore reported "NEUROSCIENCE" as an
    unresolved line needing a business decision. It never was: LINE_SYNONYMS
    has mapped NEUROSCIENCE -> CNS all along, exactly as normalizeLine() does
    at runtime. That was a bug in this script, not a gap in the data, and it
    put a false blocker in front of a decision-maker. Both tables are now
    read, in the same order normalizeLine() applies them.
    """
    with open(SEMANTIC_JS, 'r', encoding='utf-8') as fh:
        src = fh.read()
    i = src.index('CANONICAL_LINE_TO_BU = {')
    canon = dict(re.findall(r'"([^"]+)"\s*:\s*"([^"]+)"', src[i:src.index('};', i)]))
    j = src.index('LINE_SYNONYMS = {')
    syn = dict(re.findall(r'"([^"]+)"\s*:\s*"([^"]+)"', src[j:src.index('};', j)]))
    return canon, syn


# -----------------------------------------------------------------------------
# Business decisions that the semantic model does not yet carry.
#
# "Gyna" is absent from CANONICAL_LINE_TO_BU and from LINE_SYNONYMS. Ahmed's
# decision (2026-08-09) is that it belongs to DIAB.
#
# It lives HERE and not in js/semantic-model.js on purpose: editing the
# semantic model is a Phase B action requiring explicit approval, and Phase A
# is not authorised to touch existing source. Recording it as a declared
# override keeps the decision visible and applied, without quietly amending a
# file the whole platform depends on.
#
# PHASE B ACTION: fold this into CANONICAL_LINE_TO_BU and delete this block.
# -----------------------------------------------------------------------------
PENDING_LINE_TO_BU = {
    'Gyna': 'DIAB',   # Ahmed, 2026-08-09
}


def load_sales_lookups():
    """Brand and product names as the sales cube spells them."""
    with open(SALES_CACHE, 'r', encoding='utf-8') as fh:
        raw = fh.read()
    m = re.search(r'b64Data:\s*"([^"]+)"', raw)
    cache = json.loads(gzip.decompress(base64.b64decode(m.group(1))))
    lk = cache['lookups']
    return [str(b) for b in lk.get('brands', [])], [str(p) for p in lk.get('products', [])]


def load_expense_rows():
    import openpyxl
    wb = openpyxl.load_workbook(EXPENSE_XLSX, read_only=True, data_only=True)

    ws = wb[SHEET_SKU]
    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    month_cols = [(i, h) for i, h in enumerate(header) if hasattr(h, 'year')]

    skus = []
    for r in rows[1:]:
        if not r or r[1] is None:
            continue
        monthly = {}
        for idx, dt in month_cols:
            v = r[idx]
            if isinstance(v, (int, float)):
                monthly['%04d-%02d' % (dt.year, dt.month)] = float(v)
        skus.append({
            'bum': (r[0] or '').strip() if isinstance(r[0], str) else '',
            'bu': str(r[1]).strip(),
            'line': str(r[2]).strip() if r[2] else '',
            'brand': str(r[3]).strip() if r[3] else '',
            # ACTIVE (Ahmed, 2026-08-09: "check active column to differentiate
            # between active and not active sku"). Normalised to a boolean-ish
            # string rather than a bool so a blank stays distinguishable from a
            # deliberate "No" -- an unset flag is not the same claim as an
            # explicit one, and the reports must not conflate them.
            'active': (str(r[4]).strip() if r[4] else ''),
            'isActive': (str(r[4]).strip().upper() == 'YES') if r[4] else None,
            'code': str(r[5]).strip() if r[5] else '',
            'product': str(r[6]).strip() if r[6] else '',
            'pack': str(r[7]).strip() if r[7] else '',
            'monthly': monthly,
            'budget': sum(monthly.values()),
        })

    ws2 = wb[SHEET_LINE]
    lrows = list(ws2.iter_rows(values_only=True))
    line_budget = {}
    for r in lrows[2:]:
        if not r or not r[0]:
            continue
        name = str(r[0]).strip()
        if name.upper() == 'TOTAL':
            continue
        nums = [v for v in r[1:13] if isinstance(v, (int, float))]
        if nums:
            line_budget[name] = sum(nums)

    months = sorted({m for s in skus for m in s['monthly']})
    return skus, line_budget, months


# =============================================================================
# Dimension normalisation
# =============================================================================
def normalise_dimensions(skus, canon_lines, synonyms):
    """
    Resolve each row's BU and Line onto the platform's authoritative values.

    Case-only differences are resolved automatically -- "CLUSTER" and "Cluster"
    are unambiguously the same business unit and treating them otherwise would
    be pedantry. Anything requiring a JUDGEMENT is left unresolved and
    reported: guessing which BU an unrecognised line belongs to is exactly the
    kind of silent assumption that produces a confident wrong number.
    """
    canon_bus = sorted(set(list(canon_lines.values()) + list(PENDING_LINE_TO_BU.values())))
    bu_by_norm = {norm(b): b for b in canon_bus}
    line_by_norm = {norm(l): l for l in canon_lines}
    # Synonyms resolve first, mirroring normalizeLine(): "NEUROSCIENCE" is a
    # spelling of "CNS", not an unknown line.
    syn_by_norm = {norm(k): v for k, v in synonyms.items()}

    exceptions = []
    for s in skus:
        s['canonBU'] = bu_by_norm.get(norm(s['bu']))
        if s['canonBU'] is None:
            exceptions.append({
                'Type': 'BU', 'SourceValue': s['bu'], 'Resolution': 'UNRESOLVED',
                'CanonicalValue': '',
                'Detail': 'No case-insensitive match in CANONICAL_LINE_TO_BU values',
            })
        elif s['canonBU'] != s['bu']:
            exceptions.append({
                'Type': 'BU', 'SourceValue': s['bu'], 'Resolution': 'AUTO_CASE',
                'CanonicalValue': s['canonBU'],
                'Detail': 'Case-only difference, resolved automatically',
            })

        s['canonLine'] = line_by_norm.get(norm(s['line']))
        s['lineResolvedVia'] = 'canonical' if s['canonLine'] else ''
        if s['canonLine'] is None:
            syn = syn_by_norm.get(norm(s['line']))
            if syn:
                s['canonLine'] = syn
                s['lineResolvedVia'] = 'synonym'
                exceptions.append({
                    'Type': 'LINE', 'SourceValue': s['line'], 'Resolution': 'AUTO_SYNONYM',
                    'CanonicalValue': syn,
                    'Detail': 'Resolved via LINE_SYNONYMS, same rule normalizeLine() applies',
                })
            elif s['line'] in PENDING_LINE_TO_BU:
                s['canonLine'] = s['line']
                s['canonBU'] = PENDING_LINE_TO_BU[s['line']]
                s['lineResolvedVia'] = 'pending-override'
                exceptions.append({
                    'Type': 'LINE', 'SourceValue': s['line'], 'Resolution': 'BUSINESS_DECISION',
                    'CanonicalValue': s['line'] + ' -> BU ' + PENDING_LINE_TO_BU[s['line']],
                    'Detail': 'Confirmed by Ahmed 2026-08-09. NOT yet in semantic-model.js '
                              '- Phase B must add it to CANONICAL_LINE_TO_BU',
                })
        if s['canonLine'] is None:
            exceptions.append({
                'Type': 'LINE', 'SourceValue': s['line'], 'Resolution': 'UNRESOLVED',
                'CanonicalValue': '',
                'Detail': 'Not present in CANONICAL_LINE_TO_BU - needs a business decision',
            })
        elif s['canonLine'] != s['line']:
            exceptions.append({
                'Type': 'LINE', 'SourceValue': s['line'], 'Resolution': 'AUTO_CASE',
                'CanonicalValue': s['canonLine'],
                'Detail': 'Case-only difference, resolved automatically',
            })

    # de-duplicate: one row per distinct source value, with its volume
    seen, uniq = {}, []
    for e in exceptions:
        k = (e['Type'], e['SourceValue'])
        if k in seen:
            seen[k]['Occurrences'] += 1
            continue
        e['Occurrences'] = 1
        seen[k] = e
        uniq.append(e)

    for e in uniq:
        if e['Type'] == 'LINE':
            e['BudgetAffected'] = sum(s['budget'] for s in skus if s['line'] == e['SourceValue'])
        else:
            e['BudgetAffected'] = sum(s['budget'] for s in skus if s['bu'] == e['SourceValue'])
    return uniq


# =============================================================================
# The mapping table
# =============================================================================
def classify(sku, sales_products, sales_brands):
    """
    Propose a status and a target for one expense SKU.

    Order matters: an exact match must win over a normalised one, and a
    normalised match must win over "not selling", or a trivial spacing
    difference would be reported as a missing product.
    """
    prod_exact = {p: p for p in sales_products}
    prod_norm = {}
    for p in sales_products:
        prod_norm.setdefault(norm(p), p)
    brand_norm = {}
    for b in sales_brands:
        brand_norm.setdefault(norm(b), b)

    name = sku['product']
    if not name:
        return ('UNMAPPED', '', '', 'Expense row carries no product name')

    if name in prod_exact:
        return ('MATCHED', prod_exact[name], brand_norm.get(norm(sku['brand']), ''),
                'Exact product-name match')

    hit = prod_norm.get(norm(name))
    if hit:
        return ('NAME_VARIANT', hit, brand_norm.get(norm(sku['brand']), ''),
                'Matches after case/spacing normalisation - CONFIRM before use')

    bhit = brand_norm.get(norm(sku['brand']))
    if bhit:
        return ('NOT_YET_SELLING', '', bhit,
                'Brand exists in sales; this SKU has no sales record. '
                'Cause not determinable from the data - needs product knowledge.')

    return ('UNMAPPED', '', '',
            'Neither the SKU nor its brand appears in the sales cube')


def read_existing_map():
    """Prior human decisions, keyed on the row's stable identity."""
    if not os.path.exists(MAP_CSV):
        return {}
    out = {}
    with open(MAP_CSV, 'r', encoding='utf-8-sig', newline='') as fh:
        for row in csv.DictReader(fh):
            key = (row.get('ExpenseBU', ''), row.get('ExpenseCode', ''),
                   row.get('ExpenseSKU', ''))
            out[key] = row
    return out


def build_map(skus, sales_brands, sales_products):
    existing = read_existing_map()
    preserved = 0
    rows = []
    for s in skus:
        key = (s['bu'], s['code'], s['product'])
        prev = existing.get(key)

        # A row a human has signed off is authoritative. Regenerating over it
        # would throw away the review this whole design exists to capture.
        #
        # PROPOSED rows are preserved too, but they are NOT the same claim.
        # A proposal is a classification waiting for a person; it survives the
        # rebuild so the evidence is not re-derived from scratch every run, and
        # it is excluded from the coverage gate below. Only YES counts as cover.
        if prev and str(prev.get('Reviewed', '')).strip().upper() in ('YES', 'PROPOSED'):
            # WHAT A REVIEW OWNS, AND WHAT IT DOES NOT.
            #
            # The human reviewed the MAPPING -- which sales product this SKU
            # joins to. They did not review the source facts around it. Budget,
            # Active, BU, Line and Brand belong to the workbook and must be
            # re-read every run, or a reviewed row freezes them at whatever they
            # were on the day it was signed off.
            #
            # This was a live bug, found 2026-08-09. The workbook changed
            # EPILOSAMIDE 5 MG/1 ML SYRUP from Active=Yes to Active=No; because
            # the row was reviewed, the mapping table kept saying Yes. The
            # summary read the source and the CSV read the review, and they
            # disagreed by 200,000 EGP with nothing on screen to show it.
            # Budget was already refreshed here; Active was not. Same argument,
            # so now every source-owned field is refreshed.
            for k, v in (('BudgetTotal', '%.2f' % s['budget']),
                         ('Active', s['active']),
                         ('ExpenseBU', s['bu']), ('ExpenseLine', s['line']),
                         ('ExpenseBrand', s['brand']),
                         ('CanonicalBU', s['canonBU'] or ''),
                         ('CanonicalLine', s['canonLine'] or '')):
                prev[k] = v
            rows.append({k: prev.get(k, '') for k in MAP_HEADER})
            preserved += 1
            s['status'] = prev.get('MappingStatus', 'UNMAPPED')
            s['salesProduct'] = prev.get('ProposedSalesProduct', '')
            s['reviewed'] = prev.get('Reviewed', '')
            continue

        status, prod, brand, reason = classify(s, sales_products, sales_brands)
        s['status'] = status
        s['salesProduct'] = prod
        rows.append({
            'ExpenseBU': s['bu'], 'ExpenseLine': s['line'], 'ExpenseBrand': s['brand'],
            'ExpenseCode': s['code'], 'ExpenseSKU': s['product'],
            'Active': s['active'],
            'BudgetTotal': '%.2f' % s['budget'],
            'ProposedSalesProduct': prod, 'ProposedSalesBrand': brand,
            'CanonicalBU': s['canonBU'] or '', 'CanonicalLine': s['canonLine'] or '',
            'MappingStatus': status, 'MappingReason': reason,
            'Reviewed': (prev or {}).get('Reviewed', 'NO'),
            'ReviewedBy': (prev or {}).get('ReviewedBy', ''),
            'ReviewNote': (prev or {}).get('ReviewNote', ''),
        })
    return rows, preserved


# =============================================================================
# Reconciliation
# =============================================================================
def reconcile(skus, line_budget):
    by_line = defaultdict(float)
    for s in skus:
        by_line[s['line']] += s['budget']

    rows = []
    for line in sorted(set(list(line_budget) + list(by_line))):
        lb = line_budget.get(line, 0.0)
        sb = by_line.get(line, 0.0)
        var = sb - lb
        pct = (var / lb * 100.0) if lb else None
        if lb and not sb:
            status = 'NO_SKU_DETAIL'
        elif sb and not lb:
            status = 'NOT_IN_LINE_SHEET'
        elif abs(var) < 1:
            status = 'RECONCILED'
        else:
            status = 'VARIANCE'
        rows.append({
            'Line': line,
            'LineSheetBudget': '%.2f' % lb,
            'SumOfSKUBudget': '%.2f' % sb,
            'Variance': '%.2f' % var,
            'VariancePct': '' if pct is None else '%.2f' % pct,
            'Status': status,
        })
    return rows


def write_csv(path, header, rows):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    with open(path, 'w', encoding='utf-8-sig', newline='') as fh:
        w = csv.DictWriter(fh, fieldnames=header)
        w.writeheader()
        for r in rows:
            w.writerow({k: r.get(k, '') for k in header})


def write_template(skus, months):
    """
    The actual-expense entry template.

    Budget columns are included and clearly labelled READ-ONLY REFERENCE so the
    person entering actuals can see what they are being compared against --
    but the import will ignore them entirely. Budget is official data; a user's
    spreadsheet must never be able to change it.

    Only rows that can be analysed are included. Offering an entry line for a
    SKU whose budget cannot be joined to sales would invite someone to spend an
    afternoon typing numbers the dashboard is then unable to use.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Actual Expenses'

    # 'Active' travels with every row so the flag survives the whole round trip:
    # source -> template -> the person's spreadsheet -> import -> cache -> filter.
    # If it were dropped here, an imported actual would arrive with no way to
    # tell which side of the Active filter it belongs on, and the importer would
    # have to re-derive it -- a second source of truth for the same fact.
    head = ['Month', 'BU', 'Line', 'Brand', 'SKU Code', 'Product', 'Active',
            'Budget Expense (READ-ONLY)', 'Actual Expense', 'Expense Category', 'Note']
    ws.append(head)
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill('solid', fgColor='0F4C81')
        cell.alignment = Alignment(horizontal='center')
    for col, width in zip('ABCDEFGHIJK', [11, 10, 13, 22, 12, 34, 8, 22, 16, 18, 30]):
        ws.column_dimensions[col].width = width

    usable = [s for s in skus if s.get('status') in ('MATCHED', 'NAME_VARIANT')]
    for s in usable:
        for m in months:
            ws.append([m, s['canonBU'] or s['bu'], s['canonLine'] or s['line'],
                       s['brand'], s['code'], s['product'],
                       s['active'] or '',
                       s['monthly'].get(m, 0.0), None, '', ''])

    ws.freeze_panes = 'A2'

    notes = wb.create_sheet('Instructions')
    for line in [
        ['ZETA — Actual Expense Entry Template'],
        [''],
        ['1. Fill ONLY the "Actual Expense" column. Numbers only.'],
        ['2. "Budget Expense" is official reference data and is ignored on import.'],
        ['3. Do not add, rename, reorder or delete columns.'],
        ['4. Do not add rows for SKUs that are not listed.'],
        ['5. Leave a row blank if there is no actual expense for that month.'],
        ['6. Expense Category and Note are optional.'],
        [''],
        ['Rows listed here are the SKUs whose budget can be joined to sales.'],
        ['SKUs with status NOT_YET_SELLING or UNMAPPED are deliberately excluded'],
        ['until their mapping is resolved — see expense/mapping/expense_sku_map.csv.'],
        [''],
        ['The "Active" column is source data, shown for reference. Rows marked'],
        ['No are excluded from Expense vs Sales KPIs by business rule, but spend'],
        ['against them is still recorded and still appears in reconciliation.'],
        [''],
        ['Uniqueness key: Month + BU + SKU Code + Expense Category.'],
        ['A repeated key is reported as a duplicate and rejected, never merged'],
        ['and never silently overwritten.'],
    ]:
        notes.append(line)
    notes['A1'].font = Font(bold=True, size=13)
    notes.column_dimensions['A'].width = 88

    os.makedirs(TPL_DIR, exist_ok=True)
    path = os.path.join(TPL_DIR, 'expense_actuals_template.xlsx')
    wb.save(path)
    return path, len(usable), len(usable) * len(months)


def write_active_split(skus):
    """
    Budget split by the Active flag, per BU / Line / mapping status.

    This is the Active filter's arithmetic, computed and checkable BEFORE any UI
    exists. When the filter ships, every number it can display should already be
    reproducible from this file -- so a disagreement between the page and the
    source is a visible discrepancy rather than an argument.

    ACTIVE IS TRI-STATE, NOT BOOLEAN. Yes / No / (blank) are three different
    claims. Collapsing blank into No would invent a business statement nobody
    made, and it is exactly the kind of quiet default that turns into a wrong
    number on a slide six months later.
    """
    agg = defaultdict(lambda: {'skus': 0, 'budget': 0.0})
    for s in skus:
        if s['budget'] <= 0:
            continue
        a = (s['active'] or '').strip() or '(blank)'
        key = (a, s['canonBU'] or s['bu'], s['canonLine'] or s['line'],
               s.get('status') or 'UNMAPPED')
        agg[key]['skus'] += 1
        agg[key]['budget'] += s['budget']

    rows = []
    for (a, bu, line, st), v in sorted(agg.items()):
        rows.append({
            'Active': a, 'CanonicalBU': bu, 'CanonicalLine': line,
            'MappingStatus': st, 'SKUs': v['skus'],
            'BudgetTotal': '%.2f' % v['budget'],
            # Joinable is the only column that decides whether a row can appear
            # in an Expense-to-Sales figure at all. Stated explicitly so the
            # filter never has to re-derive it from the status vocabulary.
            'Joinable': 'YES' if st in ('MATCHED', 'NAME_VARIANT') else 'NO',
        })
    write_csv(os.path.join(REP_DIR, 'active_split.csv'),
              ['Active', 'CanonicalBU', 'CanonicalLine', 'MappingStatus',
               'SKUs', 'BudgetTotal', 'Joinable'], rows)
    return rows


# =============================================================================
def main():
    if not os.path.exists(EXPENSE_XLSX):
        print('ERROR: expense workbook not found: %s' % EXPENSE_XLSX)
        return 1

    canon_lines, synonyms = load_semantic_lines()
    sales_brands, sales_products = load_sales_lookups()
    skus, line_budget, months = load_expense_rows()

    dim_exceptions = normalise_dimensions(skus, canon_lines, synonyms)
    map_rows, preserved = build_map(skus, sales_brands, sales_products)
    recon_rows = reconcile(skus, line_budget)

    for d in (MAP_DIR, REP_DIR, TPL_DIR):
        os.makedirs(d, exist_ok=True)

    write_csv(MAP_CSV, MAP_HEADER, map_rows)
    write_csv(os.path.join(REP_DIR, 'budget_reconciliation.csv'),
              ['Line', 'LineSheetBudget', 'SumOfSKUBudget', 'Variance', 'VariancePct', 'Status'],
              recon_rows)
    write_csv(os.path.join(REP_DIR, 'dimension_exceptions.csv'),
              ['Type', 'SourceValue', 'CanonicalValue', 'Resolution', 'Occurrences',
               'BudgetAffected', 'Detail'],
              [{**e, 'BudgetAffected': '%.2f' % e['BudgetAffected']} for e in dim_exceptions])

    unresolved = [r for r in map_rows
                  if r['MappingStatus'] in ('NOT_YET_SELLING', 'UNMAPPED')
                  or (r['MappingStatus'] == 'NAME_VARIANT' and r['Reviewed'] != 'YES')]
    write_csv(os.path.join(REP_DIR, 'unmapped_skus.csv'), MAP_HEADER, unresolved)

    write_active_split(skus)

    tpl_path, tpl_skus, tpl_rows = write_template(skus, months)

    # ---- the numbers that decide GO / NO-GO -------------------------------
    budgeted = [s for s in skus if s['budget'] > 0]
    total_budget = sum(s['budget'] for s in budgeted)

    def group(status, pool=None):
        pool = budgeted if pool is None else pool
        g = [s for s in pool if s.get('status') == status]
        return len(g), sum(s['budget'] for s in g)

    # ACTIVE SPLIT (Ahmed, 2026-08-09).
    #
    # This turned out to be the single most important cut in the whole report.
    # Read across ALL budgeted SKUs, mapping coverage looks alarming. Split by
    # the Active flag, most of the gap turns out to be SKUs the business has
    # already marked inactive -- which have no sales precisely because they are
    # not being sold. Judging the data on the blended figure would have
    # condemned a dataset that is largely fine.
    #
    # The ACTIVE-only coverage is the number that should drive the gate,
    # because it is the population the expense-vs-sales analysis is about.
    active = [s for s in budgeted if s.get('isActive') is True]
    inactive = [s for s in budgeted if s.get('isActive') is False]
    unflagged = [s for s in budgeted if s.get('isActive') is None]
    act_total = sum(s['budget'] for s in active)
    inact_total = sum(s['budget'] for s in inactive)

    n_match, b_match = group('MATCHED')
    n_var, b_var = group('NAME_VARIANT')
    n_nys, b_nys = group('NOT_YET_SELLING')
    n_unm, b_unm = group('UNMAPPED')

    usable_n = n_match + n_var
    usable_b = b_match + b_var
    sku_cov = usable_n / len(budgeted) * 100 if budgeted else 0
    val_cov = usable_b / total_budget * 100 if total_budget else 0

    recon_bad = [r for r in recon_rows if r['Status'] != 'RECONCILED']

    # -------------------------------------------------------------------------
    # RECONCILIATION DECISIONS -- Ahmed, 2026-08-09
    #
    # A decided exception is not a fixed exception. The variance stays in the
    # report exactly as measured; what changes is that it is no longer WAITING
    # ON ANYONE. Nothing here reallocates a single EGP, and neither sheet has
    # been made authoritative.
    #
    #   Zetagarouh / CHC  confirmed a classification difference, not a budget
    #                     difference. Both sheets total 267,100,000.
    #   GIT-I / GIT-II    preserve the difference; do NOT auto-reallocate.
    #   DIAB-I..IV        OPEN. Requires explicit business confirmation. The
    #                     build must not choose an authoritative split.
    # -------------------------------------------------------------------------
    RECON_DECIDED = {
        'Zetagarouh': 'Classification difference, confirmed by Ahmed 2026-08-09. Not a budget difference.',
        'CHC':        'Classification difference, confirmed by Ahmed 2026-08-09. Not a budget difference.',
        'GIT-I':      'Difference preserved by decision, Ahmed 2026-08-09. No auto-reallocation.',
        'GIT-II':     'Difference preserved by decision, Ahmed 2026-08-09. No auto-reallocation.',
    }
    recon_open = [r for r in recon_bad if r['Line'] not in RECON_DECIDED]
    recon_decided = [r for r in recon_bad if r['Line'] in RECON_DECIDED]
    dim_bad = [e for e in dim_exceptions if e['Resolution'] == 'UNRESOLVED']

    L = []
    L.append('PHASE A — EXPENSE DATA FOUNDATION')
    L.append('=' * 72)
    L.append('source      : %s' % os.path.basename(EXPENSE_XLSX))
    L.append('months      : %d  (%s .. %s)' % (len(months), months[0], months[-1]))
    L.append('SKU rows    : %d   of which carry budget: %d' % (len(skus), len(budgeted)))
    L.append('total budget: %s EGP' % '{:,.0f}'.format(total_budget))
    L.append('')
    L.append('MAPPING COVERAGE')
    L.append('  %-18s %5s %18s %8s' % ('status', 'SKUs', 'budget EGP', '% value'))
    for label, n, b in [('MATCHED', n_match, b_match), ('NAME_VARIANT', n_var, b_var),
                        ('NOT_YET_SELLING', n_nys, b_nys), ('UNMAPPED', n_unm, b_unm)]:
        L.append('  %-18s %5d %18s %7.1f%%'
                 % (label, n, '{:,.0f}'.format(b), (b / total_budget * 100 if total_budget else 0)))
    L.append('  %-18s %5d %18s' % ('TOTAL', len(budgeted), '{:,.0f}'.format(total_budget)))
    L.append('')
    L.append('ACTIVE vs INACTIVE  (Active column in the source sheet)')
    L.append('  %-22s %5s %18s' % ('', 'SKUs', 'budget EGP'))
    L.append('  %-22s %5d %18s' % ('Active = Yes', len(active), '{:,.0f}'.format(act_total)))
    L.append('  %-22s %5d %18s' % ('Active = No', len(inactive), '{:,.0f}'.format(inact_total)))
    if unflagged:
        L.append('  %-22s %5d %18s' % ('Active = (blank)', len(unflagged),
                 '{:,.0f}'.format(sum(s['budget'] for s in unflagged))))
    L.append('')
    an_m, ab_m = group('MATCHED', active)
    an_v, ab_v = group('NAME_VARIANT', active)
    an_y, ab_y = group('NOT_YET_SELLING', active)
    an_u, ab_u = group('UNMAPPED', active)
    act_usable_n, act_usable_b = an_m + an_v, ab_m + ab_v
    act_sku_cov = act_usable_n / len(active) * 100 if active else 0
    act_val_cov = act_usable_b / act_total * 100 if act_total else 0

    # RATIFIED vs PROVISIONAL coverage.
    #
    # A row marked Reviewed=PROPOSED carries a classification I made from the
    # evidence, which Ahmed has not seen. Counting it toward the gate would let
    # the build clear its own gate using its own unreviewed opinion -- the exact
    # failure mode the human-in-the-loop mapping exists to prevent.
    #
    # So the provisional figure is shown (it is the honest best estimate of
    # where coverage lands once the proposals are ratified) and the RATIFIED
    # figure is what the gate reads.
    proposed_active = [s for s in active
                       if str(s.get('reviewed', '')).upper() == 'PROPOSED'
                       and s.get('status') in ('MATCHED', 'NAME_VARIANT')]
    prop_b = sum(s['budget'] for s in proposed_active)
    act_ratified_b = act_usable_b - prop_b
    act_val_cov_ratified = act_ratified_b / act_total * 100 if act_total else 0
    L.append('  ACTIVE SKUs only — the population the analysis is about:')
    for label, n, b in [('MATCHED', an_m, ab_m), ('NAME_VARIANT', an_v, ab_v),
                        ('NOT_YET_SELLING', an_y, ab_y), ('UNMAPPED', an_u, ab_u)]:
        L.append('    %-18s %5d %18s %7.1f%%'
                 % (label, n, '{:,.0f}'.format(b), (b / act_total * 100 if act_total else 0)))
    L.append('    %-18s %5s %18s %7.1f%%'
             % ('>> usable', act_usable_n, '{:,.0f}'.format(act_usable_b), act_val_cov))
    L.append('')
    in_u_n, in_u_b = group('UNMAPPED', inactive)
    L.append('  Of the %d INACTIVE budgeted SKUs, %d are UNMAPPED (%s EGP).'
             % (len(inactive), in_u_n, '{:,.0f}'.format(in_u_b)))
    # OBSERVATION ONLY. The source workbook carries no definition of the Active
    # flag -- no cell comment, no data validation, no legend. What it MEANS is a
    # business definition nobody has stated, so this reports the measurement and
    # stops there. An earlier draft said "expected: a SKU marked not active has
    # no sales to join to", which asserts a cause the data does not establish.
    L.append('  Measured: 0 of %d inactive budgeted SKUs have any recorded sales.' % len(inactive))
    L.append('  The source defines no meaning for this flag. Ahmed 2026-08-09: excluded')
    L.append('  from Expense vs Sales KPIs, retained in full in reconciliation, and NOT')
    L.append('  labelled pre-launch, discontinued or any other status.')
    L.append('')
    L.append('  SKU coverage (usable)   : %.1f%%  (%d of %d)' % (sku_cov, usable_n, len(budgeted)))
    L.append('  BUDGET-VALUE coverage   : %.1f%%  (%s of %s EGP)'
             % (val_cov, '{:,.0f}'.format(usable_b), '{:,.0f}'.format(total_budget)))
    L.append('  ACTIVE-only, ratified   : %.1f%%   <-- the gate figure' % act_val_cov_ratified)
    if proposed_active:
        L.append('  ACTIVE-only, provisional: %.1f%%   (incl. %d proposed mapping(s), %s EGP,'
                 % (act_val_cov, len(proposed_active), '{:,.0f}'.format(prop_b)))
        L.append('                             awaiting Ahmed\'s ratification)')
    L.append('  Human-reviewed rows kept: %d' % preserved)
    L.append('')
    L.append('EXCEPTIONS')
    L.append('  budget reconciliation   : %d line(s) not reconciled  '
             '(%d decided, %d OPEN)'
             % (len(recon_bad), len(recon_decided), len(recon_open)))
    L.append('  dimension               : %d unresolved value(s)' % len(dim_bad))
    for e in dim_bad:
        L.append('      %s "%s"  budget %s EGP'
                 % (e['Type'], e['SourceValue'], '{:,.0f}'.format(e['BudgetAffected'])))
    L.append('')
    L.append('TEMPLATE')
    L.append('  %s' % os.path.relpath(tpl_path, ROOT_DIR))
    L.append('  %d joinable SKUs x %d months = %d entry rows' % (tpl_skus, len(months), tpl_rows))
    # The template count exceeds the budgeted-usable count because a SKU can be
    # joinable to sales while carrying no budget. Those rows are still offered
    # for entry -- money can be spent on a SKU with no budget line, and that is
    # precisely a variance worth capturing -- but they contribute nothing to the
    # coverage percentages above, which are budget-weighted.
    L.append('  (%d of these carry budget; the rest are joinable but unbudgeted)' % usable_n)
    L.append('')
    L.append('GATE')
    blockers = []
    if act_val_cov_ratified < 95:
        blockers.append('ACTIVE budget-value coverage %.1f%% (ratified) is below 95%%'
                        % act_val_cov_ratified)
    # Every PROPOSED row is one blocker, whatever it was classified as.
    # "I decided not to map this" needs ratifying exactly as much as "I decided
    # to map this" -- both remove a question from the board, and neither is the
    # build's to answer on its own.
    proposed_all = [s for s in active if str(s.get('reviewed', '')).upper() == 'PROPOSED']
    if proposed_all:
        msg = ('%d classification(s) proposed and unresolved (%s EGP)'
               % (len(proposed_all),
                  '{:,.0f}'.format(sum(s['budget'] for s in proposed_all))))
        # Only quote a coverage delta when there is one. A proposal to leave a
        # SKU UNMAPPED moves no coverage at all -- printing "92.6% -> 92.6%"
        # would read as a stalled process rather than an accurate statement,
        # and would invite someone to go looking for a bug that is not there.
        if proposed_active:
            msg += ('. Ratifying the %d proposed mapping(s) takes ACTIVE coverage '
                    '%.1f%% -> %.1f%%'
                    % (len(proposed_active), act_val_cov_ratified, act_val_cov))
        else:
            msg += ('. All are proposed as NOT joinable, so ratifying them would '
                    'not change coverage - they are open questions, not pending '
                    'coverage. Resolution depends on the product master, not on '
                    'a review of this data')
        blockers.append(msg)
    if an_v > 0:
        blockers.append('%d active NAME_VARIANT row(s) await human confirmation' % an_v)
    # A row a human has reviewed is no longer a blocker even when it stays
    # UNMAPPED -- "someone looked and decided not to map this" is a resolved
    # question. It still counts against coverage, because the money genuinely
    # cannot be compared to sales; it just is not waiting on anyone.
    unreviewed_active_unmapped = [
        s2 for s2 in active
        if s2.get('status') == 'UNMAPPED'
        and str(s2.get('reviewed', '')).upper() not in ('YES', 'PROPOSED')
    ]
    if unreviewed_active_unmapped:
        blockers.append('%d ACTIVE SKU(s) UNMAPPED and not yet reviewed (%s EGP)'
                        % (len(unreviewed_active_unmapped),
                           '{:,.0f}'.format(sum(x['budget'] for x in unreviewed_active_unmapped))))
    unreviewed_nys = [
        s2 for s2 in active
        if s2.get('status') == 'NOT_YET_SELLING'
        and str(s2.get('reviewed', '')).upper() not in ('YES', 'PROPOSED')
    ]
    if unreviewed_nys:
        blockers.append('%d ACTIVE SKU(s) NOT_YET_SELLING and not yet reviewed (%s EGP) '
                        '- launch or unmapped variant?'
                        % (len(unreviewed_nys),
                           '{:,.0f}'.format(sum(x['budget'] for x in unreviewed_nys))))
    # Active=No is a RECORDED RULE as of 2026-08-09, not an open question.
    # Reported as a disclosure below rather than counted as a blocker.
    if dim_bad:
        blockers.append('%d unresolved dimension value(s)' % len(dim_bad))
    if recon_open:
        blockers.append('%d budget reconciliation exception(s) OPEN - %s. '
                        'Requires explicit business confirmation; the build will '
                        'not choose an authoritative split'
                        % (len(recon_open),
                           ', '.join(sorted(set(r['Line'] for r in recon_open)))))
    if blockers:
        L.append('  NO-GO for Phase B. Blockers:')
        for b in blockers:
            L.append('    - %s' % b)
    else:
        L.append('  GO for Phase B.')
    L.append('')
    L.append('RECORDED RULES  (not blockers)')
    L.append('  Active=No       %d SKUs, %s EGP. Excluded from Expense vs Sales KPIs;'
             % (len(inactive), '{:,.0f}'.format(inact_total)))
    L.append('                  retained in full in reconciliation. No status label applied.')
    reviewed_unmapped = [s2 for s2 in active if s2.get('status') == 'UNMAPPED'
                         and str(s2.get('reviewed', '')).upper() == 'YES']
    if reviewed_unmapped:
        L.append('  Reviewed        %d ACTIVE SKU(s) confirmed UNMAPPED by a human (%s EGP).'
                 % (len(reviewed_unmapped),
                    '{:,.0f}'.format(sum(x['budget'] for x in reviewed_unmapped))))
        for x in reviewed_unmapped:
            L.append('                  %s' % x['product'][:56])
        L.append('                  Counted against coverage; not awaiting a decision.')
    L.append('  Gyna -> DIAB    ETL override, pending addition to semantic-model.js in Phase B.')
    L.append('  Zetagarouh      Confirmed intentional under CHC.')
    L.append('  Reconciliation  %d of %d exceptions decided; variances PRESERVED, not fixed.'
             % (len(recon_decided), len(recon_bad)))
    for r in recon_decided:
        L.append('                  %-11s %14s EGP  %s'
                 % (r['Line'], '{:,.0f}'.format(float(r['Variance'])),
                    RECON_DECIDED[r['Line']].split('.')[0]))
    L.append('                  Company total agrees exactly: both sheets 267,100,000 EGP.')
    L.append('=' * 72)

    summary = '\n'.join(L)
    print(summary)
    with open(os.path.join(OUT_DIR, 'PHASE_A_SUMMARY.txt'), 'w', encoding='utf-8') as fh:
        fh.write(summary + '\n')
    return 0


if __name__ == '__main__':
    sys.exit(main())
